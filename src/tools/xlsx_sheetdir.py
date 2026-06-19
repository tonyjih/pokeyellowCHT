#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
xlsx_sheetdir.py - round-trip src/xlsx/*.xlsx <-> per-sheet TSV/CSV folders.

Default layout:
  src/xlsx/core.xlsx
    -> src/xlsx_tsv/core/workbook.json
    -> src/xlsx_tsv/core/000__Main.tsv
    -> src/xlsx_tsv/core/001__SomeSheet.tsv

The TSV files are the review/edit source of truth.
The xlsx files are regenerated for the existing pokeyellowCHS import tools.

Notes:
- TSV is the default because game text frequently contains punctuation and quotes.
- A small workbook.json manifest preserves sheet order, sheet names, dimensions,
  and cell types so numeric columns remain numeric when rebuilt.
- Import can use an existing xlsx as template to preserve styles/column widths.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import shutil
import sys
import tempfile
import unicodedata
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.cell.cell import Cell
except ModuleNotFoundError as exc:  # pragma: no cover
    raise SystemExit(
        "error: openpyxl is required. Install in WSL with: sudo apt install python3-openpyxl"
    ) from exc

VERSION = "xlsx-sheetdir-roundtrip-2026-06-19-v1"
INVALID_FILENAME_CHARS_RE = re.compile(r'[<>:"/\\|?*\x00-\x1F]')
CONTROL_CHARS_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


@dataclass(frozen=True)
class FormatSpec:
    name: str
    delimiter: str
    extension: str


FORMATS: Dict[str, FormatSpec] = {
    "tsv": FormatSpec("tsv", "\t", "tsv"),
    "csv": FormatSpec("csv", ",", "csv"),
}


def eprint(*args: object) -> None:
    print(*args, file=sys.stderr)


def resolve_repo_path(repo: Path, path: str | Path) -> Path:
    p = Path(path)
    return p if p.is_absolute() else repo / p


def read_xlsx_list(repo: Path, xlsx_list: Path) -> List[Path]:
    if not xlsx_list.exists():
        raise SystemExit(f"error: xlsx list not found: {xlsx_list}")
    out: List[Path] = []
    base_dir = xlsx_list.parent
    for raw in xlsx_list.read_text(encoding="utf-8", errors="replace").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        p = Path(line)
        if not p.is_absolute():
            # TomJinW/pokeyellowCHS xlsxList.txt usually contains just data.xlsx.
            p = base_dir / p
        out.append(p.resolve())
    return out


def sanitize_filename(name: str, fallback: str = "Sheet") -> str:
    s = unicodedata.normalize("NFKC", name)
    s = INVALID_FILENAME_CHARS_RE.sub("_", s)
    s = s.strip().strip(".")
    if not s:
        s = fallback
    # Keep filenames readable but avoid absurd path lengths on Windows.
    if len(s) > 80:
        s = s[:80].rstrip()
    return s


def unique_path_name(base: str, used: set[str]) -> str:
    name = base
    i = 2
    while name.lower() in used:
        name = f"{base}_{i}"
        i += 1
    used.add(name.lower())
    return name


def cell_type_and_value(cell: Cell) -> Tuple[str, str]:
    """Return (type, text) for manifest + TSV."""
    v = cell.value
    if v is None:
        return "blank", ""

    # openpyxl data_type values include: s, n, b, d, f, e, inlineStr, str.
    dt = cell.data_type

    if dt == "f":
        return "formula", str(v)
    if dt == "b":
        return "bool", "TRUE" if bool(v) else "FALSE"
    if dt == "n":
        if isinstance(v, bool):
            return "bool", "TRUE" if v else "FALSE"
        if isinstance(v, int):
            return "int", str(v)
        if isinstance(v, float):
            if math.isfinite(v) and v.is_integer():
                return "int", str(int(v))
            return "float", repr(v)
        return "number", str(v)
    if dt == "d" or isinstance(v, (datetime, date, time)):
        if isinstance(v, (datetime, date, time)):
            return "datetime", v.isoformat()
        return "datetime", str(v)
    if dt == "e":
        return "error", str(v)

    # Strings and anything else become text.
    return "str", str(v)


def parse_value(raw: str, cell_type: str) -> Any:
    if cell_type == "blank":
        return None if raw == "" else raw
    if cell_type == "str":
        return raw
    if cell_type == "formula":
        return raw
    if cell_type in {"int", "number"}:
        if raw == "":
            return None
        try:
            return int(raw)
        except ValueError:
            try:
                return float(raw)
            except ValueError:
                return raw
    if cell_type == "float":
        if raw == "":
            return None
        try:
            return float(raw)
        except ValueError:
            return raw
    if cell_type == "bool":
        return raw.strip().upper() in {"TRUE", "1", "YES", "Y"}
    if cell_type == "datetime":
        if raw == "":
            return None
        # Keep it conservative. If parsing fails, preserve text.
        for parser in (datetime.fromisoformat, date.fromisoformat, time.fromisoformat):
            try:
                return parser(raw)  # type: ignore[arg-type]
            except ValueError:
                pass
        return raw
    if cell_type == "error":
        return raw
    # New cells without metadata: infer lightly.
    if raw == "":
        return None
    return raw


def copy_sheet_format(src_ws: Any, dst_ws: Any) -> None:
    """Copy common visual metadata from template sheet into destination sheet."""
    dst_ws.sheet_view = copy(src_ws.sheet_view)
    dst_ws.sheet_format = copy(src_ws.sheet_format)
    dst_ws.sheet_properties = copy(src_ws.sheet_properties)
    dst_ws.page_margins = copy(src_ws.page_margins)
    dst_ws.page_setup = copy(src_ws.page_setup)
    dst_ws.print_options = copy(src_ws.print_options)
    dst_ws.freeze_panes = src_ws.freeze_panes

    for key, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[key] = copy(dim)
    for key, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[key] = copy(dim)
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))


def clear_worksheet_values(ws: Any, max_row: int, max_col: int) -> None:
    rows = max(ws.max_row or 1, max_row, 1)
    cols = max(ws.max_column or 1, max_col, 1)
    for row in ws.iter_rows(min_row=1, max_row=rows, min_col=1, max_col=cols):
        for cell in row:
            cell.value = None


def write_tsv(path: Path, rows: List[List[str]], fmt: FormatSpec, encoding: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding=encoding, newline="") as f:
        writer = csv.writer(f, delimiter=fmt.delimiter, lineterminator="\n", quoting=csv.QUOTE_MINIMAL)
        writer.writerows(rows)


def read_tsv(path: Path, fmt: FormatSpec, encoding: str) -> List[List[str]]:
    with path.open("r", encoding=encoding, newline="") as f:
        reader = csv.reader(f, delimiter=fmt.delimiter)
        return [list(row) for row in reader]


def export_workbook(xlsx_path: Path, out_root: Path, fmt: FormatSpec, encoding: str, overwrite: bool) -> Path:
    if not xlsx_path.exists():
        raise SystemExit(f"error: xlsx not found: {xlsx_path}")

    book_dir_name = sanitize_filename(xlsx_path.stem, fallback="workbook")
    book_dir = out_root / book_dir_name
    if book_dir.exists() and overwrite:
        shutil.rmtree(book_dir)
    book_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xlsx_path, data_only=False)
    used: set[str] = set()
    sheets_meta: List[Dict[str, Any]] = []

    for idx, ws in enumerate(wb.worksheets):
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        safe_sheet = sanitize_filename(ws.title, fallback=f"Sheet{idx:03d}")
        file_stem = unique_path_name(f"{idx:03d}__{safe_sheet}", used)
        rel_file = f"{file_stem}.{fmt.extension}"
        sheet_path = book_dir / rel_file

        rows: List[List[str]] = []
        types: Dict[str, str] = {}
        number_formats: Dict[str, str] = {}

        for r in range(1, max_row + 1):
            out_row: List[str] = []
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                typ, txt = cell_type_and_value(cell)
                out_row.append(txt)
                if typ != "blank":
                    key = f"{r},{c}"
                    types[key] = typ
                    if cell.number_format and cell.number_format != "General":
                        number_formats[key] = cell.number_format
            rows.append(out_row)

        write_tsv(sheet_path, rows, fmt, encoding)
        sheets_meta.append({
            "index": idx,
            "name": ws.title,
            "file": rel_file,
            "max_row": max_row,
            "max_col": max_col,
            "types": types,
            "number_formats": number_formats,
        })

    manifest = {
        "version": VERSION,
        "format": fmt.name,
        "delimiter": "\\t" if fmt.delimiter == "\t" else fmt.delimiter,
        "encoding": encoding,
        "source_xlsx": str(xlsx_path),
        "source_relpath": xlsx_path.name,
        "workbook_stem": xlsx_path.stem,
        "sheets": sheets_meta,
    }
    (book_dir / "workbook.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return book_dir


def load_manifest(book_dir: Path) -> Dict[str, Any]:
    manifest_path = book_dir / "workbook.json"
    if not manifest_path.exists():
        raise SystemExit(f"error: manifest not found: {manifest_path}")
    return json.loads(manifest_path.read_text(encoding="utf-8"))


def import_workbook(book_dir: Path, output_xlsx: Path, template_xlsx: Optional[Path], encoding_override: Optional[str], remove_extra_sheets: bool) -> Path:
    manifest = load_manifest(book_dir)
    fmt_name = manifest.get("format", "tsv")
    fmt = FORMATS.get(fmt_name)
    if fmt is None:
        raise SystemExit(f"error: unsupported format in manifest: {fmt_name}")
    encoding = encoding_override or manifest.get("encoding", "utf-8")

    if template_xlsx and template_xlsx.exists():
        wb = load_workbook(template_xlsx, data_only=False)
    else:
        wb = Workbook()
        # Remove default sheet. It will be recreated below.
        while wb.worksheets:
            wb.remove(wb.worksheets[0])

    desired_names = [s["name"] for s in manifest.get("sheets", [])]

    if remove_extra_sheets:
        for ws in list(wb.worksheets):
            if ws.title not in desired_names:
                wb.remove(ws)

    # Reorder by recreating only when necessary is complex; instead keep template sheets and
    # set workbook._sheets at the end. openpyxl stores worksheets in wb._sheets.
    ordered_sheets = []

    for sheet_meta in manifest.get("sheets", []):
        name = sheet_meta["name"]
        rel_file = sheet_meta["file"]
        sheet_path = book_dir / rel_file
        rows = read_tsv(sheet_path, fmt, encoding)
        max_row = max(int(sheet_meta.get("max_row", 0) or 0), len(rows), 1)
        max_col = max(int(sheet_meta.get("max_col", 0) or 0), max((len(r) for r in rows), default=0), 1)
        types: Dict[str, str] = dict(sheet_meta.get("types", {}))
        number_formats: Dict[str, str] = dict(sheet_meta.get("number_formats", {}))

        if name in wb.sheetnames:
            ws = wb[name]
            clear_worksheet_values(ws, max_row=max_row, max_col=max_col)
        else:
            ws = wb.create_sheet(title=name)

        for r_idx in range(1, max_row + 1):
            row_vals = rows[r_idx - 1] if r_idx - 1 < len(rows) else []
            for c_idx in range(1, max_col + 1):
                raw = row_vals[c_idx - 1] if c_idx - 1 < len(row_vals) else ""
                typ = types.get(f"{r_idx},{c_idx}")
                if typ is None:
                    # New cell. Keep empty empty; otherwise text by default to avoid
                    # accidentally converting game strings like 1F or 01.
                    value = None if raw == "" else raw
                else:
                    value = parse_value(raw, typ)
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                nf = number_formats.get(f"{r_idx},{c_idx}")
                if nf:
                    cell.number_format = nf

        ordered_sheets.append(ws)

    # Preserve any non-listed sheets after listed sheets if not removing extras.
    if not remove_extra_sheets:
        listed = {ws.title for ws in ordered_sheets}
        for ws in wb.worksheets:
            if ws.title not in listed:
                ordered_sheets.append(ws)

    if ordered_sheets:
        wb._sheets = ordered_sheets  # type: ignore[attr-defined]

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    # Save atomically to reduce risk if Excel has file locked or save fails.
    with tempfile.NamedTemporaryFile(prefix=output_xlsx.name + ".", suffix=".tmp", dir=str(output_xlsx.parent), delete=False) as tf:
        tmp_path = Path(tf.name)
    try:
        wb.save(tmp_path)
        tmp_path.replace(output_xlsx)
    finally:
        if tmp_path.exists():
            tmp_path.unlink(missing_ok=True)
    return output_xlsx


def find_book_dirs(tsv_root: Path) -> List[Path]:
    if not tsv_root.exists():
        raise SystemExit(f"error: tsv root not found: {tsv_root}")
    out = []
    for p in sorted(tsv_root.iterdir()):
        if p.is_dir() and (p / "workbook.json").exists():
            out.append(p)
    return out


def command_export(args: argparse.Namespace) -> int:
    repo = Path(args.repo).resolve()
    fmt = FORMATS[args.format]
    out_root = resolve_repo_path(repo, args.out_root)
    xlsx_path = resolve_repo_path(repo, args.xlsx)
    book_dir = export_workbook(xlsx_path, out_root, fmt, args.encoding, args.overwrite)
    print(f"exported: {xlsx_path} -> {book_dir}")
    return 0


def command_export_all(args: argparse.Namespace) -> int:
    repo = Path(args.repo).resolve()
    fmt = FORMATS[args.format]
    out_root = resolve_repo_path(repo, args.out_root)
    xlsx_list = resolve_repo_path(repo, args.xlsx_list)
    xlsx_files = read_xlsx_list(repo, xlsx_list)
    for xlsx_path in xlsx_files:
        book_dir = export_workbook(xlsx_path, out_root, fmt, args.encoding, args.overwrite)
        print(f"exported: {xlsx_path} -> {book_dir}")
    print(f"done: {len(xlsx_files)} workbook(s)")
    return 0


def command_import(args: argparse.Namespace) -> int:
    repo = Path(args.repo).resolve()
    book_dir = resolve_repo_path(repo, args.book_dir)
    output = resolve_repo_path(repo, args.output)
    template = resolve_repo_path(repo, args.template) if args.template else None
    out = import_workbook(book_dir, output, template, args.encoding, args.remove_extra_sheets)
    print(f"imported: {book_dir} -> {out}")
    return 0


def command_import_all(args: argparse.Namespace) -> int:
    repo = Path(args.repo).resolve()
    tsv_root = resolve_repo_path(repo, args.tsv_root)
    out_dir = resolve_repo_path(repo, args.out_dir)
    template_dir = resolve_repo_path(repo, args.template_dir) if args.template_dir else out_dir

    if args.xlsx_list:
        xlsx_list = resolve_repo_path(repo, args.xlsx_list)
        stems = {p.stem for p in read_xlsx_list(repo, xlsx_list)}
        book_dirs = [p for p in find_book_dirs(tsv_root) if p.name in stems]
    else:
        book_dirs = find_book_dirs(tsv_root)

    for book_dir in book_dirs:
        manifest = load_manifest(book_dir)
        stem = manifest.get("workbook_stem", book_dir.name)
        output = out_dir / f"{stem}.xlsx"
        template = template_dir / f"{stem}.xlsx" if template_dir else None
        out = import_workbook(book_dir, output, template, args.encoding, args.remove_extra_sheets)
        print(f"imported: {book_dir} -> {out}")
    print(f"done: {len(book_dirs)} workbook(s)")
    return 0


def build_argparser() -> argparse.ArgumentParser:
    ap = argparse.ArgumentParser(description="Round-trip xlsx workbooks to per-sheet TSV/CSV folders.")
    ap.add_argument("--version", action="version", version=f"%(prog)s {VERSION}")
    sub = ap.add_subparsers(dest="cmd", required=True)

    common_export = argparse.ArgumentParser(add_help=False)
    common_export.add_argument("--repo", default=".")
    common_export.add_argument("--out-root", default="src/xlsx_tsv")
    common_export.add_argument("--format", choices=sorted(FORMATS), default="tsv")
    common_export.add_argument("--encoding", default="utf-8")
    common_export.add_argument("--overwrite", action="store_true", help="Remove existing output workbook folder before export.")

    p = sub.add_parser("export", parents=[common_export], help="Export one xlsx to a sheet folder.")
    p.add_argument("--xlsx", required=True)
    p.set_defaults(func=command_export)

    p = sub.add_parser("export-all", parents=[common_export], help="Export all workbooks listed by xlsxList.txt.")
    p.add_argument("--xlsx-list", default="src/xlsx/xlsxList.txt")
    p.set_defaults(func=command_export_all)

    p = sub.add_parser("import", help="Rebuild one xlsx from one workbook folder.")
    p.add_argument("--repo", default=".")
    p.add_argument("--book-dir", required=True, help="Example: src/xlsx_tsv/core")
    p.add_argument("--output", required=True, help="Example: src/xlsx/core.xlsx")
    p.add_argument("--template", default=None, help="Optional xlsx template to preserve styles. Defaults to none.")
    p.add_argument("--encoding", default=None)
    p.add_argument("--remove-extra-sheets", action="store_true", default=True, help="Remove sheets not listed in workbook.json. Default: true.")
    p.set_defaults(func=command_import)

    p = sub.add_parser("import-all", help="Rebuild all xlsx files from sheet folders.")
    p.add_argument("--repo", default=".")
    p.add_argument("--tsv-root", default="src/xlsx_tsv")
    p.add_argument("--out-dir", default="src/xlsx")
    p.add_argument("--template-dir", default="src/xlsx", help="Use existing xlsx files as style templates. Default: src/xlsx")
    p.add_argument("--xlsx-list", default="src/xlsx/xlsxList.txt")
    p.add_argument("--encoding", default=None)
    p.add_argument("--remove-extra-sheets", action="store_true", default=True, help="Remove sheets not listed in workbook.json. Default: true.")
    p.set_defaults(func=command_import_all)

    return ap


def main(argv: Optional[Sequence[str]] = None) -> int:
    ap = build_argparser()
    args = ap.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
