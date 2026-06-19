#!/usr/bin/env python3
"""
Build text-driven Traditional Chinese assets for tonyjih/pokeyellowCHT.

V14 adds a safer private-slot allocator. Old IME Chinese candidate codepoints are reserved, default/special symbols are protected, and new CHT-only glyphs are allocated from DFS physical holes that were not present in the original charmap.

This is the formal pipeline after the DFS smoke tests:

1. Extract the text that the current build pipeline imports from src/xlsx/*.xlsx.
2. Convert that text to Traditional Chinese with OpenCC.
3. Collect the final CHT glyph character set, while preserving existing special non-CJK codepoints such as %/kana.
4. Allocate DFS codepoints from the existing DFS slot pool, using a lock file for stability.
5. Generate candidate charmap.txt while preserving original non-CJK/special entries; keep RGBDS charmap.asm unchanged for the default byte-emitting pipeline.
6. Render CHT glyphs from Fusion Pixel Font and write patched ChineseFonts_*.bin copies.

V14 keeps compatible-slot priority, but no longer steals existing non-CJK/symbol entries such as ˉ=$0105 for private CHT glyphs. It uses unused DFS physical holes first; optionally, --allow-old-non-ime-spare can fall back to old non-IME CJK slots.

By default this script is SAFE: it writes everything to reports/cht_assets and does not modify src/.
Use --apply-assets and/or --apply-xlsx only after reviewing the generated preview and reports.

Expected to live at src/tools/build_cht_text_assets.py and to be run from the repo root:

    python src/tools/build_cht_text_assets.py --repo .

Dependencies:

    python -m pip install openpyxl opencc-python-reimplemented pillow

It also expects src/tools/build_dfs_font_from_fusion.py v5 or later to exist, because that
script already knows how to download Fusion Pixel Font and pack DFS 12x12 glyphs.
"""

from __future__ import annotations

import argparse
import collections
import csv
import dataclasses
import importlib.util
import re
import shlex
import shutil
import sys
import time
from pathlib import Path
from types import SimpleNamespace
from typing import Dict, Iterable, Iterator, List, Optional, Sequence, Tuple

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit("Missing dependency: openpyxl. Install with: python -m pip install openpyxl") from exc

try:
    from opencc import OpenCC
except ImportError:  # pragma: no cover
    OpenCC = None  # type: ignore

VERSION = "cht-assets-2026-06-18-v15"

CJK_RE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff]")
TOKEN_RE = re.compile(r"(<[^>]+>|\[[^\]]+\]|#[A-Za-z0-9_]*|[@&+^~]|\\.|\$[0-9A-Fa-f]+)")
CHARMAP_TXT_RE = re.compile(r"^(.+?)=\$?([0-9A-Fa-f]{4})\s*(?:;.*)?$")
CHARMAP_ASM_CJK_RE = re.compile(r'^\s*charmap\s+"(.+?)"\s*,', re.IGNORECASE)

# These legacy special glyphs are known to have game-specific meaning or are
# commonly used by menus/text. V14 keeps every existing non-CJK/symbol entry as
# a real occupied slot; new CHT-only glyphs are placed in DFS physical holes
# that were not present in the original charmap.
ALWAYS_PRESERVE_SPECIAL_CHARS = set("îñ%　、。·—～…‘’“”「」『』《》！？，．－：；（）／０１２３４５６７８９ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ")

# DFS physical pages that exist in src/dfs/ChineseFonts_XX_[HL].bin through FontAB.
# The lookup uses the low 6 bits of the first DFS byte. Pages mapped to FF do not
# have backing font storage. This mirrors build_dfs_font_from_fusion.py.
DFS_FONTAB = [
    "FF", "01", "02", "03", "04", "05", "06", "07", "08", "09", "0A", "0B", "0C", "0D", "0E", "0F",
    "10", "11", "12", "13", "FF", "FF", "FF", "FF", "18", "19", "1A", "1B", "1C", "1D", "1E", "1F",
    "FF", "FF", "FF", "FF", "FF", "FF", "FF", "FF", "28", "29", "2A", "2B", "2C", "2D", "2E", "FF",
]
DFS_SUPPORTED_HIS = [hi for hi, file_code in enumerate(DFS_FONTAB) if file_code != "FF"]
# Original charmap rows deliberately never use these low bytes. Keep following
# that convention so generated two-byte text cannot collide with common control
# / terminator / token-ish byte values.
DFS_UNSAFE_LOW_BYTES = {0x00, 0x14, 0x15, 0x16, 0x22, 0x3F, *range(0x4B, 0x60), 0xFD, 0xFE, 0xFF}
DFS_SAFE_LOW_BYTES = [lo for lo in range(0x01, 0xFD) if lo not in DFS_UNSAFE_LOW_BYTES]


@dataclasses.dataclass(frozen=True)
class TextLine:
    file: str
    sheet: str
    row: int
    col: int
    original: str
    converted: str  # visible CHT text used for reports / charmap extraction
    encoded: str    # text after import-token compression, used for charmap/glyph audit
    xlsx_write: str # exact value to write back to xlsx; excludes synthetic importer suffixes


@dataclasses.dataclass(frozen=True)
class BuildTask:
    tool: str
    xlsx_rel: str
    mode: int
    ver: str
    build_mode_arg: str
    raw_args: Tuple[str, ...]
    raw_command: str


@dataclasses.dataclass(frozen=True)
class OriginalEntry:
    char: str
    code: int
    line_no: int

    @property
    def code_hex(self) -> str:
        return f"${self.code:04X}"


@dataclasses.dataclass(frozen=True)
class Allocation:
    char: str
    code: int
    source: str
    note: str = ""

    @property
    def code_hex(self) -> str:
        return f"${self.code:04X}"


@dataclasses.dataclass(frozen=True)
class LockEntry:
    char: str
    code: int
    source: str
    note: str = ""


@dataclasses.dataclass
class ExtractResult:
    lines: List[TextLine]
    used_chs: collections.Counter[str]
    used_cht: collections.Counter[str]


def is_cjk_char(ch: str) -> bool:
    return len(ch) == 1 and bool(CJK_RE.fullmatch(ch))


def cjk_chars(text: str) -> List[str]:
    return CJK_RE.findall(text)


def is_glyph_char(ch: str) -> bool:
    """Return True for non-ASCII visible characters that need DFS/charmap coverage.

    This includes CJK, kana, and fullwidth punctuation. The previous versions only
    counted CJK; data.xlsx also contains Japanese kana strings such as をんゥェ1,
    which still pass through charmap.replaceText and therefore must be present.
    """
    if len(ch) != 1:
        return False
    code = ord(ch)
    if code < 0x80:
        return False
    if ch.isspace():
        return False
    # Do not treat Private Use placeholders created by protect_tokens as real glyphs.
    if 0xE000 <= code <= 0xF8FF:
        return False
    return True


def glyph_chars(text: str) -> List[str]:
    # Remove protected/control tokens first, then collect every visible non-ASCII glyph.
    protected, mapping = protect_tokens(text or "")
    for token in mapping:
        protected = protected.replace(token, "")
    return [ch for ch in protected if is_glyph_char(ch)]


class IdentityOpenCC:
    def convert(self, text: str) -> str:
        return text


def make_opencc(mode: str):
    if mode.lower() in {"none", "identity", "raw", "no-opencc"}:
        return IdentityOpenCC()
    if OpenCC is None:
        raise SystemExit(
            "Missing dependency: opencc-python-reimplemented. Install with: "
            "python -m pip install opencc-python-reimplemented"
        )
    try:
        return OpenCC(mode)
    except Exception:
        if not mode.endswith(".json"):
            return OpenCC(mode + ".json")
        raise


def protect_tokens(text: str) -> Tuple[str, Dict[str, str]]:
    mapping: Dict[str, str] = {}

    def repl(match: re.Match[str]) -> str:
        token = f"\ue000{len(mapping)}\ue001"
        mapping[token] = match.group(0)
        return token

    return TOKEN_RE.sub(repl, text), mapping


def restore_tokens(text: str, mapping: Dict[str, str]) -> str:
    for token, original in mapping.items():
        text = text.replace(token, original)
    return text


def convert_text(text: str, cc: OpenCC) -> str:
    protected, mapping = protect_tokens(text)
    return restore_tokens(cc.convert(protected), mapping)


def read_xlsx_list(repo: Path, include_lgpe: bool = False) -> List[Path]:
    xlsx_dir = repo / "src" / "xlsx"
    list_path = xlsx_dir / "xlsxList.txt"
    if list_path.exists():
        names = [line.strip() for line in list_path.read_text(encoding="utf-8", errors="replace").splitlines()]
        names = [name for name in names if name and not name.startswith("#")]
        paths = [xlsx_dir / name for name in names]
    else:
        paths = sorted(xlsx_dir.glob("*.xlsx"))

    if include_lgpe:
        lgpe = xlsx_dir / "LGPE.xlsx"
        if lgpe.exists() and lgpe not in paths:
            paths.append(lgpe)

    return [p for p in paths if p.exists() and p.suffix.lower() == ".xlsx"]


def parse_prepare_tasks(repo: Path, prepare_path: Optional[Path] = None) -> List[BuildTask]:
    """Parse the python import commands from _prepare.command.

    The original build copies src/* into buildYUS and then runs commands like
    `python3 tools/_importText.py xlsx/outdoor.xlsx 5 YEUS $option`.
    This function turns those commands into typed extraction tasks, so the
    character audit follows the real build instead of guessing a workbook column.
    """
    if prepare_path is None:
        prepare_path = repo / "_prepare.command"
    if not prepare_path.exists():
        raise SystemExit(f"prepare script not found: {prepare_path}")
    text = prepare_path.read_text(encoding="utf-8", errors="replace")
    pattern = re.compile(
        r"python3\s+tools/(\S+)\s+(.+?)(?=\s+python3\s+tools/|\s+chmod\b|\s+\./_build|$)",
        re.DOTALL,
    )
    tasks: List[BuildTask] = []
    for m in pattern.finditer(text):
        tool = m.group(1).strip()
        arg_text = m.group(2).strip()
        # Remove inline comments that may appear after old disabled commands.
        arg_text = arg_text.split(" #", 1)[0].strip()
        if not arg_text:
            continue
        try:
            parts = shlex.split(arg_text, posix=False)
        except ValueError:
            parts = arg_text.split()
        if not parts or not parts[0].lower().endswith(".xlsx"):
            continue
        xlsx_rel = parts[0].replace("\\", "/")
        raw_command = f"python3 tools/{tool} {arg_text}"
        if tool in {"_importText.py", "_importText2.py"}:
            if len(parts) < 4:
                continue
            mode = int(parts[1])
            ver = parts[2]
            build_mode_arg = parts[3]
        elif tool == "_importDexEntry.py":
            if len(parts) < 5:
                continue
            mode = int(parts[1])
            build_mode_arg = parts[3]
            ver = parts[4]
        elif tool == "_importTextData.py":
            if len(parts) < 4:
                continue
            mode = int(parts[1])
            ver = parts[2]
            build_mode_arg = parts[3]
        else:
            continue
        tasks.append(BuildTask(tool, xlsx_rel, mode, ver, build_mode_arg, tuple(parts), raw_command))
    return tasks


def build_xlsx_path(repo: Path, task: BuildTask) -> Path:
    rel = task.xlsx_rel
    if rel.startswith("src/"):
        return repo / rel
    # _prepare.command runs inside buildYUS, where xlsx/... is at root.
    # In the source repo, that path lives under src/xlsx/....
    if rel.startswith("xlsx/"):
        return repo / "src" / rel
    return repo / "src" / rel


def cell_str(ws, row: int, col: int) -> str:
    value = ws.cell(row=row, column=col).value
    if value is None:
        return ""
    return str(value)


def has_ver(ver_cell: str, wanted: str) -> bool:
    return (not ver_cell) or (wanted in ver_cell)


def get_label_type(label: str) -> int:
    if not label:
        return -1
    if label.endswith(":"):
        return 0
    if label.startswith(";"):
        parts = label.split()
        if parts and parts[-1].lower() == "start":
            return 3
    return -1


def get_inst_type(inst: str) -> int:
    if not inst:
        return -1
    text_commands = ["page", "text", "line", "cont", "para", "next"]
    placer_commands = ["text_ram", "text_decimal", "text_bcd"]
    if inst in placer_commands:
        return 1
    for cmd in text_commands:
        if cmd in inst:
            return 0
    return -2


def apply_import_text_tokens(text: str) -> str:
    """Approximate the token expansion/compression that _importText.py applies before charmap.replaceText.

    This is used for character-set extraction, not for generating the final asm.
    Include both CHS and CHT phrases so the audit reflects the planned converted xlsx.
    """
    replacements = collections.OrderedDict([
        ("#MON", "#"),
        ("寶可夢", "#"),
        ("宝可梦", "#"),
        ("&", "訓練家"),
        ("+", "火箭隊"),
    ])
    out = text or ""
    for src, dst in replacements.items():
        out = out.replace(src, dst)
    return out


def add_text_line(
    lines: List[TextLine],
    used_chs: collections.Counter[str],
    used_cht: collections.Counter[str],
    cc: OpenCC,
    file_rel: str,
    sheet: str,
    row: int,
    col: int,
    original: str,
    token_mode: str = "text",
    xlsx_write_source: Optional[str] = None,
) -> None:
    if not isinstance(original, str) or original == "" or original.startswith("="):
        return
    converted_visible = convert_text(original, cc)
    if xlsx_write_source is None:
        xlsx_write = converted_visible
    else:
        xlsx_write = convert_text(xlsx_write_source, cc)
    if token_mode == "text":
        encoded_for_charmap = apply_import_text_tokens(converted_visible)
    else:
        encoded_for_charmap = converted_visible
    original_glyphs = glyph_chars(original)
    encoded_glyphs = glyph_chars(encoded_for_charmap)
    if not original_glyphs and not encoded_glyphs:
        return
    lines.append(TextLine(
        file=file_rel,
        sheet=sheet,
        row=row,
        col=col,
        original=original,
        converted=converted_visible,
        encoded=encoded_for_charmap,
        xlsx_write=xlsx_write,
    ))
    used_chs.update(original_glyphs)
    used_cht.update(encoded_glyphs)


def extract_import_text_task(repo: Path, task: BuildTask, cc: OpenCC, lines: List[TextLine], used_chs, used_cht) -> int:
    path = build_xlsx_path(repo, task)
    if not path.exists():
        raise SystemExit(f"xlsx not found for build task: {task.xlsx_rel} -> {path}")
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    count = 0
    rel = str(path.relative_to(repo))
    mode = task.mode
    for ws in wb.worksheets:
        row = 2
        max_row = min(ws.max_row or 0, 10000)
        while row <= max_row and cell_str(ws, row, mode) != "end":
            if mode - 1 > 0:
                ver_value = cell_str(ws, row, mode - 1)
                if ver_value and task.ver not in ver_value:
                    row += 1
                    continue
            label = cell_str(ws, row, mode)
            label_type = get_label_type(label)
            if label_type in {0, 3}:
                row += 1
                continue
            inst = cell_str(ws, row, mode + 1)
            content = ws.cell(row=row, column=mode + 2).value
            inst_type = get_inst_type(inst)
            if inst_type in {0, 1} and isinstance(content, str):
                before = len(lines)
                add_text_line(lines, used_chs, used_cht, cc, rel, str(ws.title), row, mode + 2, content, token_mode="text")
                count += len(lines) - before
            row += 1
    return count


def extract_dex_entry_task(repo: Path, task: BuildTask, cc: OpenCC, lines: List[TextLine], used_chs, used_cht) -> int:
    path = build_xlsx_path(repo, task)
    if not path.exists():
        raise SystemExit(f"xlsx not found for build task: {task.xlsx_rel} -> {path}")
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    count = 0
    rel = str(path.relative_to(repo))
    mode = task.mode
    # _importDexEntry.py imports category text from mode + 1.
    for ws in wb.worksheets:
        row = 2
        max_row = ws.max_row or 0
        while row <= max_row and cell_str(ws, row, 1) != "end":
            category = ws.cell(row=row, column=mode + 1).value
            if isinstance(category, str):
                before = len(lines)
                add_text_line(lines, used_chs, used_cht, cc, rel, str(ws.title), row, mode + 1, category, token_mode="raw")
                count += len(lines) - before
            row += 1
    return count


def textdata_skipped(input_ver: str, wanted: str) -> bool:
    return bool(input_ver) and wanted not in input_ver


def extract_text_data_task(repo: Path, task: BuildTask, cc: OpenCC, lines: List[TextLine], used_chs, used_cht) -> int:
    path = build_xlsx_path(repo, task)
    if not path.exists():
        raise SystemExit(f"xlsx not found for build task: {task.xlsx_rel} -> {path}")
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    count = 0
    rel = str(path.relative_to(repo))
    mode = task.mode
    for ws in wb.worksheets:
        end_char = cell_str(ws, 1, mode + 1)
        skipped = textdata_skipped(cell_str(ws, 1, mode + 2), task.ver)
        row = 2
        max_row = ws.max_row or 0
        while row <= max_row and cell_str(ws, row, mode) != "end":
            # A non-empty mode column starts a new target file block. It is not text.
            if cell_str(ws, row, mode):
                end_char = cell_str(ws, row, mode + 1)
                skipped = textdata_skipped(cell_str(ws, row, mode + 2), task.ver)
                row += 1
                continue
            replacee = cell_str(ws, row, mode + 1)
            if replacee and not skipped:
                override = ws.cell(row=row, column=mode + 3).value
                if isinstance(override, str) and override:
                    text = override
                    col = mode + 3
                else:
                    repl = ws.cell(row=row, column=mode + 2).value
                    repl_text = str(repl) if repl is not None else ""
                    text = repl_text + end_char
                    col = mode + 2
                    xlsx_write_source = repl_text
                before = len(lines)
                if 'xlsx_write_source' in locals():
                    add_text_line(lines, used_chs, used_cht, cc, rel, str(ws.title), row, col, text, token_mode="raw", xlsx_write_source=xlsx_write_source)
                    del xlsx_write_source
                else:
                    add_text_line(lines, used_chs, used_cht, cc, rel, str(ws.title), row, col, text, token_mode="raw")
                count += len(lines) - before
            row += 1
    return count


def extract_xlsx_text_from_prepare(repo: Path, mode: str, include_lgpe: bool = False) -> Tuple[ExtractResult, List[BuildTask], List[Tuple[BuildTask, int]]]:
    print(f"  loading OpenCC mode={mode} ...", flush=True)
    cc = make_opencc(mode)
    print("  parsing _prepare.command ...", flush=True)
    lines: List[TextLine] = []
    used_chs: collections.Counter[str] = collections.Counter()
    used_cht: collections.Counter[str] = collections.Counter()
    tasks = parse_prepare_tasks(repo)
    print(f"  parsed build tasks: {len(tasks)}", flush=True)
    task_counts: List[Tuple[BuildTask, int]] = []
    for task in tasks:
        print(f"  task: {task.tool} {task.xlsx_rel} mode={task.mode} ver={task.ver} ...", flush=True)
        t0 = time.time()
        if task.tool in {"_importText.py", "_importText2.py"}:
            count = extract_import_text_task(repo, task, cc, lines, used_chs, used_cht)
        elif task.tool == "_importDexEntry.py":
            count = extract_dex_entry_task(repo, task, cc, lines, used_chs, used_cht)
        elif task.tool == "_importTextData.py":
            count = extract_text_data_task(repo, task, cc, lines, used_chs, used_cht)
        else:
            count = 0
        print(f"    -> {count} text lines ({time.time() - t0:.2f}s)", flush=True)
        task_counts.append((task, count))
    if include_lgpe:
        # Optional exploratory mode; LGPE is not imported by _prepare.command.
        legacy = extract_xlsx_text_legacy(repo, 1, mode, True, True)
        lines.extend(legacy.lines)
        used_chs.update(legacy.used_chs)
        used_cht.update(legacy.used_cht)
    return ExtractResult(lines=lines, used_chs=used_chs, used_cht=used_cht), tasks, task_counts


def extract_xlsx_text_legacy(repo: Path, text_col: int, mode: str, include_lgpe: bool, all_string_cells: bool) -> ExtractResult:
    cc = make_opencc(mode)
    lines: List[TextLine] = []
    used_chs: collections.Counter[str] = collections.Counter()
    used_cht: collections.Counter[str] = collections.Counter()

    for path in read_xlsx_list(repo, include_lgpe=include_lgpe):
        rel = str(path.relative_to(repo))
        wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
        for ws in wb.worksheets:
            if all_string_cells:
                cells = ws.iter_rows()
            else:
                cells = ws.iter_rows(min_col=text_col, max_col=text_col)
            for row in cells:
                for cell in row:
                    value = cell.value
                    if not isinstance(value, str):
                        continue
                    add_text_line(lines, used_chs, used_cht, cc, rel, str(ws.title), int(cell.row), int(cell.column), value, token_mode="raw")
    return ExtractResult(lines=lines, used_chs=used_chs, used_cht=used_cht)

def parse_charmap_txt(path: Path) -> Tuple[Dict[str, OriginalEntry], Dict[int, OriginalEntry], List[str]]:
    char_to_entry: Dict[str, OriginalEntry] = {}
    code_to_entry: Dict[int, OriginalEntry] = {}
    lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    for line_no, raw in enumerate(lines, start=1):
        line = raw.strip()
        if not line or line.startswith(";"):
            continue
        m = CHARMAP_TXT_RE.match(line)
        if not m:
            continue
        char, code_hex = m.groups()
        char = char.strip()
        if len(char) != 1:
            continue
        code = int(code_hex, 16)
        entry = OriginalEntry(char=char, code=code, line_no=line_no)
        char_to_entry.setdefault(char, entry)
        code_to_entry.setdefault(code, entry)
    return char_to_entry, code_to_entry, lines




def parse_ime_reserved_codes(path: Path) -> Tuple[set[int], Dict[int, int], List[str]]:
    """Return DFS codepoints referenced by IMECharTable.asm candidate lists.

    IMECharTable stores candidates as raw two-byte DFS codepoints followed by $50.
    Those codepoints are not necessarily present in story text, but they can be
    produced by the input method and may be exchanged through player/monster names.
    Therefore V11 can reserve them so unrelated new CHT glyphs do not overwrite
    their slots.
    """
    if not path.exists():
        raise SystemExit(f"IME reserve file not found: {path}")

    label_re = re.compile(r"^\s*(IME_[A-Za-z0-9_]+_Char):")
    byte_re = re.compile(r"\$([0-9A-Fa-f]{2})")
    codes: set[int] = set()
    counts: Dict[int, int] = collections.Counter()
    labels: List[str] = []
    current_label: Optional[str] = None
    current_bytes: List[int] = []
    ended_current = False

    def flush_current() -> None:
        nonlocal current_label, current_bytes, ended_current
        if current_label is None:
            return
        labels.append(current_label)
        usable: List[int] = []
        for b in current_bytes:
            if b == 0x50:
                break
            usable.append(b)
        # Candidate bytes are hi,lo pairs. A dangling byte is ignored and reported by omission.
        for i in range(0, len(usable) - 1, 2):
            code = (usable[i] << 8) | usable[i + 1]
            codes.add(code)
            counts[code] = counts.get(code, 0) + 1
        current_label = None
        current_bytes = []
        ended_current = False

    for raw in path.read_text(encoding="utf-8", errors="replace").splitlines():
        m = label_re.match(raw)
        if m:
            flush_current()
            current_label = m.group(1)
            current_bytes = []
            ended_current = False
            continue
        if current_label is None or ended_current:
            continue
        line = raw.split(";", 1)[0]
        if "db" not in line:
            continue
        found = [int(x, 16) for x in byte_re.findall(line)]
        if not found:
            continue
        current_bytes.extend(found)
        if 0x50 in found:
            ended_current = True
            flush_current()
    flush_current()
    return codes, dict(counts), labels


def parse_ime_code_table_codes(path: Path) -> set[int]:
    """Parse DFS codepoint pairs used by IMECodeTable default symbol pages.

    IMECharTable reserves Chinese candidates. IMECodeTable also embeds default
    symbol codepoints such as punctuation/fullwidth symbols. Those should not be
    repurposed as spare CHT glyph slots or the symbol page will break.
    """
    if not path.exists():
        raise SystemExit(f"IME code table file not found: {path}")
    text = path.read_text(encoding="utf-8", errors="replace")
    byte_re = re.compile(r"\$([0-9A-Fa-f]{2})")
    codes: set[int] = set()
    for raw in text.splitlines():
        line = raw.split(";", 1)[0]
        if "db" not in line:
            continue
        vals = [int(x, 16) for x in byte_re.findall(line)]
        i = 0
        while i + 1 < len(vals):
            if vals[i] == 0x50:
                break
            code = vals[i] * 0x100 + vals[i + 1]
            codes.add(code)
            i += 2
    return codes


def code_is_compatible_with_original_char(
    ch: str,
    code: int,
    original_code_map: Dict[int, OriginalEntry],
    t2s_converter: Optional[OpenCC],
) -> Tuple[bool, str]:
    old = original_code_map.get(code)
    if old is None:
        return False, "no original char at code"
    if old.char == ch:
        return True, f"same original char {old.char}"
    simp = t2s_char(ch, t2s_converter)
    if simp and simp == old.char:
        return True, f"Traditional {ch} maps to old simplified {old.char}"
    return False, f"old char {old.char} is not compatible with {ch}"

def load_lock(path: Path) -> Dict[str, LockEntry]:
    if not path.exists():
        return {}
    result: Dict[str, LockEntry] = {}
    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            ch = (row.get("char") or "").strip()
            code_s = (row.get("code") or "").strip().replace("$", "")
            if len(ch) != 1 or not code_s:
                continue
            try:
                code = int(code_s, 16)
            except ValueError:
                continue
            result[ch] = LockEntry(
                char=ch,
                code=code,
                source=(row.get("source") or "lock").strip(),
                note=(row.get("note") or "").strip(),
            )
    return result


def write_lock(path: Path, allocations: Sequence[Allocation], old_lock: Dict[str, LockEntry]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    merged: Dict[str, Tuple[int, str, str]] = {}
    for ch, item in old_lock.items():
        merged[ch] = (item.code, item.source, item.note)
    for item in allocations:
        merged[item.char] = (item.code, item.source, item.note)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter="\t", lineterminator="\n")
        writer.writerow(["char", "code", "source", "note"])
        for ch, (code, source, note) in sorted(merged.items(), key=lambda kv: (kv[1][0], kv[0])):
            writer.writerow([ch, f"${code:04X}", source, note])


def t2s_char(ch: str, converter: Optional[OpenCC]) -> Optional[str]:
    if converter is None:
        return None
    try:
        out = converter.convert(ch)
    except Exception:
        return None
    if len(out) == 1 and out != ch:
        return out
    return None


def original_entry_is_preserved_special(entry: OriginalEntry) -> bool:
    """Return True for old charmap entries that should stay exactly as-is.

    The CHS project uses some non-CJK codepoints for special glyphs or control-like
    text placeholders. For example, % is mapped to the single-tile PP glyph.
    These entries must not be consumed by the CHT allocator, and they should not be
    re-rendered from Fusion Pixel Font as a literal percent sign. Existing kana and
    other non-CJK entries are preserved for the same reason.
    """
    return len(entry.char) == 1 and not is_cjk_char(entry.char)


def dfs_physical_hole_codes(original_code_map: Dict[int, OriginalEntry]) -> List[int]:
    """Return backed DFS codepoints not listed by the original charmap.

    These slots have physical storage in ChineseFonts_* bins, but no legacy
    charmap meaning. They are the safest place for CHT-only private glyphs such
    as 隻 when no compatible simplified slot exists.
    """
    holes: List[int] = []
    for hi in DFS_SUPPORTED_HIS:
        for lo in DFS_SAFE_LOW_BYTES:
            code = (hi << 8) | lo
            if code not in original_code_map:
                holes.append(code)
    return holes


def code_is_supported_physical_hole(code: int, original_code_map: Dict[int, OriginalEntry]) -> bool:
    hi = (code >> 8) & 0xFF
    lo = code & 0xFF
    return hi in DFS_SUPPORTED_HIS and lo in DFS_SAFE_LOW_BYTES and code not in original_code_map


def allocate_chars(
    used_cht: collections.Counter[str],
    original_char_map: Dict[str, OriginalEntry],
    original_code_map: Dict[int, OriginalEntry],
    lock: Dict[str, LockEntry],
    prefer_t2s_slots: bool,
    reserved_ime_codes: Optional[set[int]] = None,
    protected_extra_codes: Optional[set[int]] = None,
    allow_old_non_ime_spare: bool = False,
) -> Tuple[List[Allocation], List[str]]:
    preserved_special_chars = {
        ch for ch in used_cht.keys()
        if ch in original_char_map and original_entry_is_preserved_special(original_char_map[ch])
    }
    used_chars = sorted(
        (ch for ch in used_cht.keys() if ch not in preserved_special_chars),
        key=lambda ch: (-used_cht[ch], ch),
    )

    protected_extra_codes = set(protected_extra_codes or set())
    reserved_ime_codes = set(reserved_ime_codes or set())

    # V14 keeps all original non-CJK/symbol codepoints occupied. This prevents
    # collisions like 隻=$0105 where $0105 used to be the macron glyph ˉ. New
    # private CHT glyphs are placed in DFS physical holes that have backing font
    # storage but no original charmap entry.
    hard_special_chars = set(preserved_special_chars) | ALWAYS_PRESERVE_SPECIAL_CHARS
    reserved_special_codes = {
        original_char_map[ch].code
        for ch in hard_special_chars
        if ch in original_char_map and original_entry_is_preserved_special(original_char_map[ch])
    }
    # IME/default UI raw symbols are protected even if they were not parsed as a
    # normal charmap entry.
    reserved_special_codes |= set(protected_extra_codes)

    physical_hole_pool = [
        code for code in dfs_physical_hole_codes(original_code_map)
        if code not in reserved_special_codes and code not in reserved_ime_codes
    ]
    old_non_ime_spare_pool = [
        entry.code
        for entry in sorted(original_code_map.values(), key=lambda e: (e.line_no, e.code))
        if is_cjk_char(entry.char)
        and entry.code not in reserved_special_codes
        and entry.code not in reserved_ime_codes
    ]
    used_codes: set[int] = set(reserved_special_codes)
    allocations: Dict[str, Allocation] = {}
    warnings: List[str] = []
    warnings.append(
        f"protected {len(reserved_special_codes)} special/default-symbol codepoints; "
        f"found {len(physical_hole_pool)} private DFS physical holes"
    )
    if allow_old_non_ime_spare:
        warnings.append(
            f"old non-IME CJK fallback enabled: {len(old_non_ime_spare_pool)} legacy CJK slots may be repurposed if holes run out"
        )
    if preserved_special_chars:
        sample = "".join(sorted(preserved_special_chars))
        warnings.append(
            f"preserved {len(preserved_special_chars)} existing non-CJK/special charmap entries without re-rendering: {sample}"
        )
    if reserved_ime_codes:
        warnings.append(
            f"reserved {len(reserved_ime_codes)} old IME candidate codepoints; new CHT-only glyphs will not overwrite them"
        )

    # Prepare t2s once. It is used both for preferred slot selection and for
    # validating whether a locked codepoint is compatible with the old IME slot.
    t2s = None
    try:
        t2s = make_opencc("t2s")
    except BaseException:
        t2s = None
        if prefer_t2s_slots or reserved_ime_codes:
            warnings.append("OpenCC t2s unavailable; t2s compatibility checks are exact-only")

    ignored_lock_reserved_ime = 0

    # Exact original char/code preservation for CJK characters already present in the old map.
    # V12 intentionally does this BEFORE reading the lock. A stale lock from the old
    # text-only allocator may put a CHT char on a random non-IME spare slot, e.g.
    # 誌=$0508, even though the compatible old slot 志 exists. Exact/t2s-compatible
    # slots are the link-safe mapping and must win first.
    for ch in used_chars:
        if ch in allocations:
            continue
        old = original_char_map.get(ch)
        if old and old.code not in used_codes:
            source = "exact-ime" if old.code in reserved_ime_codes else "exact"
            note = f"old line {old.line_no}"
            if old.code in reserved_ime_codes:
                note += "; same old IME slot"
            allocations[ch] = Allocation(ch, old.code, source, note)
            used_codes.add(old.code)

    # Optional t2s slot preference. This is allocation only, not runtime aliasing:
    # the generated charmap will contain the Traditional char, and the glyph will be Traditional.
    if prefer_t2s_slots and t2s is not None:
        for ch in used_chars:
            if ch in allocations:
                continue
            simp = t2s_char(ch, t2s)
            if not simp:
                continue
            old = original_char_map.get(simp)
            if old and old.code not in used_codes:
                source = "t2s-slot-ime" if old.code in reserved_ime_codes else "t2s-slot"
                note = f"via {simp} old line {old.line_no}"
                if old.code in reserved_ime_codes:
                    note += "; compatible old IME slot"
                allocations[ch] = Allocation(ch, old.code, source, note)
                used_codes.add(old.code)

    # Lock is only a stability hint after exact/t2s-compatible mappings. V14
    # accepts a locked code only if it is compatible with the original char, a
    # private physical hole, or an explicitly allowed old non-IME CJK spare.
    ignored_lock_reserved_ime = 0
    ignored_lock_unsafe = 0
    for ch in used_chars:
        if ch in allocations:
            continue
        if ch in lock:
            item = lock[ch]
            if item.code in used_codes:
                warnings.append(f"lock conflict/reserved code: {ch} wants ${item.code:04X}, already used or reserved")
                continue
            old_entry = original_code_map.get(item.code)
            if old_entry is not None:
                ok, reason = code_is_compatible_with_original_char(ch, item.code, original_code_map, t2s)
                if not ok:
                    if item.code in reserved_ime_codes:
                        ignored_lock_reserved_ime += 1
                        warnings.append(f"ignored lock on IME-reserved slot: {ch} wants ${item.code:04X}; {reason}")
                        continue
                    if original_entry_is_preserved_special(old_entry):
                        ignored_lock_unsafe += 1
                        warnings.append(f"ignored lock on original special/symbol slot: {ch} wants ${item.code:04X}; old char {old_entry.char}")
                        continue
                    if not allow_old_non_ime_spare:
                        ignored_lock_unsafe += 1
                        warnings.append(f"ignored lock on old CJK non-IME slot without --allow-old-non-ime-spare: {ch} wants ${item.code:04X}; old char {old_entry.char}")
                        continue
            else:
                if not code_is_supported_physical_hole(item.code, original_code_map):
                    ignored_lock_unsafe += 1
                    warnings.append(f"ignored lock on unsupported/non-hole DFS slot: {ch} wants ${item.code:04X}")
                    continue
            if item.code in reserved_ime_codes:
                source = "lock-compatible-ime"
            elif item.code not in original_code_map:
                source = "lock-private-hole"
            else:
                source = "lock"
            allocations[ch] = Allocation(ch, item.code, source, item.note)
            used_codes.add(item.code)
    if ignored_lock_reserved_ime:
        warnings.append(f"ignored {ignored_lock_reserved_ime} incompatible lock entries on IME-reserved slots")
    if ignored_lock_unsafe:
        warnings.append(f"ignored {ignored_lock_unsafe} unsafe lock entries")

    # Allocate all remaining chars from private physical holes first. These have
    # backing DFS storage but no legacy charmap meaning, so they are safer than
    # stealing old symbol or CJK slots.
    hole_iter = (code for code in physical_hole_pool if code not in used_codes)
    for ch in used_chars:
        if ch in allocations:
            continue
        try:
            code = next(hole_iter)
        except StopIteration:
            break
        allocations[ch] = Allocation(ch, code, "new-private-hole", "allocated from DFS physical hole")
        used_codes.add(code)

    if allow_old_non_ime_spare:
        old_iter = (code for code in old_non_ime_spare_pool if code not in used_codes)
        for ch in used_chars:
            if ch in allocations:
                continue
            try:
                code = next(old_iter)
            except StopIteration:
                break
            old = original_code_map.get(code)
            note = "allocated from old non-IME CJK spare"
            if old:
                note += f"; old char {old.char} line {old.line_no}"
            allocations[ch] = Allocation(ch, code, "new-old-non-ime-spare", note)
            used_codes.add(code)

    missing = [ch for ch in used_chars if ch not in allocations]
    if missing:
        raise SystemExit(
            f"Not enough safe DFS font slots. Failed at char {missing[0]!r}; "
            f"missing={len(missing)} used_chars={len(used_chars)} private_holes={len(physical_hole_pool)}. "
            "Consider --allow-old-non-ime-spare or real DFS bank expansion."
        )

    return [allocations[ch] for ch in used_chars], warnings


def preserve_original_charmap_lines(lines: Sequence[str], allocations: Sequence[Allocation]) -> List[str]:
    """Preserve original non-CJK/special charmap lines that do not conflict.

    We remove original CJK lines because the generated CHT block is now the source
    of truth for CJK glyph slots. We also remove any original line whose code is
    reused by a generated allocation, preventing stale CHS aliases from pointing
    at newly rendered CHT glyphs. Non-CJK entries such as kana/fullwidth symbols
    are kept when they are not regenerated.
    """
    allocated_chars = {item.char for item in allocations}
    allocated_codes = {item.code for item in allocations}
    out: List[str] = []
    for raw in lines:
        line = raw.strip()
        m = CHARMAP_TXT_RE.match(line)
        if m:
            char = m.group(1).strip()
            code = int(m.group(2), 16)
            if len(char) == 1:
                if is_cjk_char(char):
                    continue
                if char in allocated_chars or code in allocated_codes:
                    continue
        out.append(raw)
    return out


def write_charmap_txt(path: Path, original_lines: Sequence[str], allocations: Sequence[Allocation]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    preserved = preserve_original_charmap_lines(original_lines, allocations)
    with path.open("w", encoding="utf-8", newline="") as f:
        for line in preserved:
            f.write(line + "\n")
        if preserved and preserved[-1].strip():
            f.write("\n")
        f.write("; Auto-generated CHT DFS charmap block. Do not edit by hand.\n")
        for item in sorted(allocations, key=lambda a: (a.code, a.char)):
            f.write(f"{item.char}=${item.code:04X}\n")


def rgbds_escape_char(ch: str) -> str:
    if ch == "\\":
        return "\\\\"
    if ch == '"':
        return '\\"'
    return ch


def write_charmap_asm(path: Path, repo: Path, allocations: Sequence[Allocation]) -> None:
    """Copy the original charmap.asm unchanged.

    Important: the default build path uses Python import tools to convert CJK text
    to explicit byte sequences via src/charmap.txt. Vanilla RGBDS 0.7 charmap
    entries are 8-bit, so emitting lines such as:

        charmap "寶", $05, $2E

    will fail with `syntax error, unexpected ","`. Therefore, this tool must not
    generate wide CJK RGBDS charmap entries. The report still keeps allocation
    data in charmap_allocation.tsv and src/charmap.txt.
    """
    src = repo / "src" / "charmap.asm"
    path.parent.mkdir(parents=True, exist_ok=True)
    if src.exists():
        shutil.copy2(src, path)
    else:
        path.write_text('; charmap.asm was not found; default CHT pipeline does not generate it.\n', encoding="utf-8")
    (path.parent / "charmap_asm_note.txt").write_text(
        "charmap.asm was intentionally left unchanged.\n"
        "The default CHS/CHT pipeline emits CJK bytes through Python using charmap.txt.\n"
        "Do not add generated lines like: charmap \"寶\", $05, $2E; vanilla RGBDS 0.7 rejects them.\n",
        encoding="utf-8",
    )


def write_text_reports(
    out: Path,
    result: ExtractResult,
    allocations: Sequence[Allocation],
    warnings: Sequence[str],
    tasks: Sequence[BuildTask] = (),
    task_counts: Sequence[Tuple[BuildTask, int]] = (),
) -> None:
    out.mkdir(parents=True, exist_ok=True)

    (out / "used_chars_chs.txt").write_text("".join(sorted(result.used_chs.keys())) + "\n", encoding="utf-8")
    (out / "used_chars_cht.txt").write_text("".join(sorted(result.used_cht.keys())) + "\n", encoding="utf-8")
    (out / "font_glyph_chars.txt").write_text("".join(a.char for a in allocations) + "\n", encoding="utf-8")

    with (out / "extracted_text_lines.tsv").open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter="\t", lineterminator="\n")
        writer.writerow(["file", "sheet", "row", "col", "original", "converted_visible", "xlsx_write", "encoded_for_charmap"])
        for line in result.lines:
            writer.writerow([line.file, line.sheet, line.row, line.col, line.original, line.converted, line.xlsx_write, line.encoded])

    if tasks:
        with (out / "build_text_manifest.tsv").open("w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter="\t", lineterminator="\n")
            writer.writerow(["tool", "xlsx", "mode", "ver", "build_mode_arg", "raw_args", "raw_command"] )
            for task in tasks:
                writer.writerow([task.tool, task.xlsx_rel, task.mode, task.ver, task.build_mode_arg, " ".join(task.raw_args), task.raw_command])
    if task_counts:
        with (out / "extract_task_summary.tsv").open("w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter="\t", lineterminator="\n")
            writer.writerow(["tool", "xlsx", "mode", "ver", "text_lines"] )
            for task, count in task_counts:
                writer.writerow([task.tool, task.xlsx_rel, task.mode, task.ver, count])

    with (out / "char_frequency_cht.tsv").open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter="\t", lineterminator="\n")
        writer.writerow(["char", "count"])
        for ch, count in result.used_cht.most_common():
            writer.writerow([ch, count])

    with (out / "charmap_allocation.tsv").open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter="\t", lineterminator="\n")
        writer.writerow(["char", "code", "source", "note"])
        for item in sorted(allocations, key=lambda a: (a.code, a.char)):
            writer.writerow([item.char, item.code_hex, item.source, item.note])

    source_counts = collections.Counter(item.source for item in allocations)
    summary_lines = [
        f"build_cht_text_assets {VERSION}",
        f"generated={time.strftime('%Y-%m-%d %H:%M:%S')}",
        f"text_lines={len(result.lines)}",
        f"unique_chs_glyph_chars={len(result.used_chs)}",
        f"unique_cht_glyph_chars={len(result.used_cht)}",
        f"allocations={len(allocations)}",
        f"build_tasks={len(tasks)}",
        "",
        "allocation_sources:",
    ]
    for source, count in sorted(source_counts.items()):
        summary_lines.append(f"  {source}: {count}")
    if warnings:
        summary_lines.append("")
        summary_lines.append("warnings:")
        for warning in warnings:
            summary_lines.append(f"  - {warning}")
    (out / "summary.txt").write_text("\n".join(summary_lines) + "\n", encoding="utf-8")


def apply_assets(repo: Path, out: Path, apply_charmap_asm: bool) -> None:
    if apply_charmap_asm:
        raise SystemExit("--apply-charmap-asm is disabled in v6: charmap.asm must stay unchanged; apply src/charmap.txt and DFS bins only.")
    src_dir = repo / "src"
    dfs_src = src_dir / "dfs"
    backup_dir = repo / "backups" / f"cht_assets_{time.strftime('%Y%m%d_%H%M%S')}"
    backup_dir.mkdir(parents=True, exist_ok=True)

    for rel in ["src/charmap.txt"]:
        src_path = repo / rel
        if src_path.exists():
            dest = backup_dir / rel
            dest.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(src_path, dest)

    shutil.copy2(out / "charmap.txt", src_dir / "charmap.txt")
    dfs_out = out / "dfs"
    if dfs_out.exists():
        for src_file in dfs_out.glob("ChineseFonts_*_[HL].bin"):
            target = dfs_src / src_file.name
            if target.exists():
                dest = backup_dir / "src" / "dfs" / src_file.name
                dest.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(target, dest)
            shutil.copy2(src_file, target)

    print(f"applied assets; backups written to {backup_dir}")


def apply_xlsx_conversion(repo: Path, result: ExtractResult) -> None:
    # Group by workbook to avoid opening files repeatedly.
    backup_dir = repo / "backups" / f"cht_xlsx_{time.strftime('%Y%m%d_%H%M%S')}"
    grouped: Dict[str, List[TextLine]] = collections.defaultdict(list)
    for line in result.lines:
        if line.original != line.converted or line.converted != line.xlsx_write:
            grouped[line.file].append(line)
    if not grouped:
        print("xlsx conversion: no changed cells")
        return

    backup_dir.mkdir(parents=True, exist_ok=True)
    for rel, lines in grouped.items():
        path = repo / rel
        dest = backup_dir / rel
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(path, dest)
        wb = openpyxl.load_workbook(path)
        for item in lines:
            ws = wb[item.sheet]
            ws.cell(row=item.row, column=item.col).value = item.xlsx_write
        wb.save(path)
        print(f"converted {rel}: {len(lines)} cells")
    print(f"xlsx backups written to {backup_dir}")


def import_font_helper(repo: Path):
    candidate_paths = [
        repo / "src" / "tools" / "build_dfs_font_from_fusion.py",
        Path(__file__).resolve().with_name("build_dfs_font_from_fusion.py"),
    ]
    for path in candidate_paths:
        if path.exists():
            spec = importlib.util.spec_from_file_location("build_dfs_font_from_fusion_helper", path)
            if spec is None or spec.loader is None:
                continue
            mod = importlib.util.module_from_spec(spec)
            # Python 3.12 dataclasses expects the module object to exist in
            # sys.modules while class decorators are executed.
            sys.modules[spec.name] = mod
            try:
                spec.loader.exec_module(mod)  # type: ignore[union-attr]
            except Exception:
                sys.modules.pop(spec.name, None)
                raise
            return mod
    raise SystemExit(
        "Could not find src/tools/build_dfs_font_from_fusion.py. "
        "Copy the v5 autodl script there first."
    )


def generate_font_assets(
    repo: Path,
    out: Path,
    allocations: Sequence[Allocation],
    args: argparse.Namespace,
) -> Tuple[Path, object]:
    helper = import_font_helper(repo)
    font_args = SimpleNamespace(
        font_path=args.font_path,
        font_release=args.font_release,
        font_url=args.font_url,
        font_cache=args.font_cache,
        fusion_pixel_size=args.fusion_pixel_size,
        fusion_width=args.fusion_width,
        fusion_lang=args.fusion_lang,
        fusion_format=args.fusion_format,
        force_download=args.force_download,
    )
    font_source = helper.resolve_font_path(font_args, repo)
    font = helper.ImageFont.truetype(str(font_source.font_path), args.font_size)

    dfs_out = out / "dfs"
    patcher = helper.DFSFontPatcher(repo, dfs_out, False)
    patcher.prepare()

    rendered = []
    for item in allocations:
        entry = helper.CharmapEntry(item.char, item.code, 0)
        glyph = helper.render_char_12x12(font, item.char, args.threshold, args.x_offset, args.y_offset, args.invert)
        patcher.patch_glyph(entry, glyph)
        rendered.append((item.char, entry, glyph))
    patcher.save()

    preview = Path(args.preview)
    if not preview.is_absolute():
        preview = repo / preview
    helper.draw_preview(rendered, preview, args.cols, args.scale)

    font_report = out / "font_report.txt"
    lines = [
        "CHT DFS font generation report",
        f"font_path={font_source.font_path}",
        f"release_tag={font_source.release_tag}",
        f"asset_name={font_source.asset_name}",
        f"sha256={font_source.sha256}",
        f"glyphs={len(rendered)}",
        f"preview={preview}",
        "",
        "touched_font_files:",
    ]
    for path in sorted(patcher.touched):
        try:
            lines.append(str(path.relative_to(repo)))
        except ValueError:
            lines.append(str(path))
    font_report.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return preview, font_source


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Build text-driven CHT charmap and DFS font assets.")
    parser.add_argument("--version", action="version", version=f"%(prog)s {VERSION}")
    parser.add_argument("--repo", default=".", help="Repo root. Default: current directory.")
    parser.add_argument("--out", default="reports/cht_assets", help="Output directory. Default: reports/cht_assets")
    parser.add_argument("--mode", default="s2twp", help="OpenCC mode. Default: s2twp. Use 'none' after manual xlsx edits to audit/generate assets from the exact current workbook text without re-normalizing it.")
    parser.add_argument(
        "--xlsx-build-col-index",
        type=int,
        default=5,
        help=(
            "0-based xlsx column index used by the original _prepare.command/import scripts. "
            "Default: 5, which corresponds to Excel column F."
        ),
    )
    parser.add_argument(
        "--text-col",
        type=int,
        default=None,
        help=(
            "Override with a 1-based Excel column number. If omitted, "
            "--xlsx-build-col-index + 1 is used."
        ),
    )
    parser.add_argument("--include-lgpe", action="store_true", help="Also read src/xlsx/LGPE.xlsx if present. This is exploratory; _prepare.command does not import it.")
    parser.add_argument("--legacy-column-scan", action="store_true", help="Old v2 behavior: ignore _prepare.command and scan --text-col / --all-string-cells.")
    parser.add_argument("--all-string-cells", action="store_true", help="With --legacy-column-scan, audit every string cell instead of only --text-col")
    parser.add_argument("--lock", default="src/dfs/charmap_cht.lock.tsv", help="Stable allocation lock file")
    parser.add_argument("--ignore-lock", action="store_true", help="Ignore the existing lock and write a fresh lock from this run. Recommended when switching to IME-compatible allocation from older text-only allocations.")
    parser.add_argument("--slot-charmap", default=None, help="Charmap.txt to use as the full DFS slot pool. Use the original CHS charmap for iterative --mode none runs after generated charmap.txt has been applied.")
    parser.add_argument("--reserve-ime", default=None, help="Path to original/base src/dfs/IMECharTable.asm. Codepoints used by this IME are reserved from unrelated new allocations, while exact/t2s-compatible replacements may still use their original slots.")
    parser.add_argument("--reserve-ime-code-table", default=None, help="Path to original/base src/dfs/IMECodeTable.asm. DFS symbol codepoints embedded in IMEDefault are protected so the symbol page keeps working.")
    parser.add_argument("--no-prefer-t2s-slots", action="store_true", help="Do not use old simplified slots as allocation hints")
    parser.add_argument("--allow-old-non-ime-spare", action="store_true", help="If private DFS holes are insufficient, allow repurposing old CJK slots that are not in the original IME candidate table. Default is off to maximize communication/IME compatibility.")
    parser.add_argument("--no-font", action="store_true", help="Skip Fusion font generation; only write charmap/reports")
    parser.add_argument("--apply-assets", action="store_true", help="Copy generated charmap.txt and DFS bins into src/. Safe backups are created.")
    parser.add_argument("--apply-charmap-asm", action="store_true", help="Disabled in v6. charmap.asm is not generated for vanilla RGBDS; leave it unchanged.")
    parser.add_argument("--apply-xlsx", action="store_true", help="Convert selected xlsx cells in place. Safe backups are created.")

    # Font options mirrored from build_dfs_font_from_fusion.py.
    parser.add_argument("--font-path", help="Local TTF/OTF path. If omitted, Fusion Pixel Font is downloaded.")
    parser.add_argument("--font-release", default="latest", help="Fusion Pixel Font release tag, or 'latest'. Default: latest.")
    parser.add_argument("--font-url", help="Explicit font/archive URL override")
    parser.add_argument("--font-cache", default=".cache/fusion-pixel-font", help="Font download cache dir")
    parser.add_argument("--fusion-pixel-size", type=int, default=12, help="Fusion Pixel release pixel size. Default: 12")
    parser.add_argument("--fusion-width", default="monospaced", choices=["monospaced", "proportional"], help="Fusion Pixel width mode")
    parser.add_argument("--fusion-lang", default="zh_hant", help="Fusion Pixel language-specific glyph version")
    parser.add_argument("--fusion-format", default="ttf", choices=["ttf", "otf"], help="Font format")
    parser.add_argument("--force-download", action="store_true", help="Re-download cached font asset")
    parser.add_argument("--font-size", type=int, default=12, help="Pillow render font size. Default: 12")
    parser.add_argument("--threshold", type=int, default=128, help="Ink threshold. Default: 128")
    parser.add_argument("--x-offset", type=int, default=0, help="Glyph X offset")
    parser.add_argument("--y-offset", type=int, default=0, help="Glyph Y offset")
    parser.add_argument("--invert", action="store_true", help="Invert font rendering threshold")
    parser.add_argument("--preview", default="reports/cht_assets/dfs_font_preview.png", help="Preview PNG path")
    parser.add_argument("--cols", type=int, default=32, help="Preview columns")
    parser.add_argument("--scale", type=int, default=4, help="Preview scale")

    args = parser.parse_args(argv)

    repo = Path(args.repo).resolve()
    out = Path(args.out)
    if not out.is_absolute():
        out = repo / out
    out.mkdir(parents=True, exist_ok=True)

    charmap_path = repo / "src" / "charmap.txt"
    if not charmap_path.exists():
        raise SystemExit(f"charmap not found: {charmap_path}")
    slot_charmap_path = Path(args.slot_charmap) if args.slot_charmap else charmap_path
    if not slot_charmap_path.is_absolute():
        slot_charmap_path = repo / slot_charmap_path
    if not slot_charmap_path.exists():
        raise SystemExit(f"slot charmap not found: {slot_charmap_path}")
    if slot_charmap_path != charmap_path:
        print(f"using DFS slot pool from: {slot_charmap_path}", flush=True)

    print("extracting xlsx text...")
    if str(args.mode).lower() in {"none", "identity", "raw", "no-opencc"}:
        print("  OpenCC disabled: using exact current xlsx text for report/charmap/assets.", flush=True)
    tasks: List[BuildTask] = []
    task_counts: List[Tuple[BuildTask, int]] = []
    if args.legacy_column_scan:
        text_col = args.text_col if args.text_col is not None else args.xlsx_build_col_index + 1
        if args.all_string_cells:
            print("  source: legacy all string cells")
        else:
            print(f"  source: legacy Excel column {text_col} (original build col index {args.xlsx_build_col_index})")
        result = extract_xlsx_text_legacy(repo, text_col, args.mode, args.include_lgpe, args.all_string_cells)
    else:
        result, tasks, task_counts = extract_xlsx_text_from_prepare(repo, args.mode, args.include_lgpe)
        print(f"  source: parsed _prepare.command build tasks: {len(tasks)}")
        for task, count in task_counts:
            print(f"    {task.tool} {task.xlsx_rel} mode={task.mode} ver={task.ver}: {count} text lines")
    print(f"  text lines: {len(result.lines)}")
    print(f"  unique CHT glyph chars: {len(result.used_cht)}")

    original_char_map, original_code_map, original_lines = parse_charmap_txt(slot_charmap_path)

    reserved_ime_codes: set[int] = set()
    reserved_ime_counts: Dict[int, int] = {}
    reserved_ime_labels: List[str] = []
    if args.reserve_ime:
        reserve_ime_path = Path(args.reserve_ime)
        if not reserve_ime_path.is_absolute():
            reserve_ime_path = repo / reserve_ime_path
        reserved_ime_codes, reserved_ime_counts, reserved_ime_labels = parse_ime_reserved_codes(reserve_ime_path)
        print(
            f"using IME reserve set from: {reserve_ime_path} "
            f"({len(reserved_ime_codes)} unique codepoints, {sum(reserved_ime_counts.values())} candidate refs)",
            flush=True,
        )

    protected_extra_codes: set[int] = set()
    if args.reserve_ime_code_table:
        reserve_code_path = Path(args.reserve_ime_code_table)
        if not reserve_code_path.is_absolute():
            reserve_code_path = repo / reserve_code_path
        protected_extra_codes |= parse_ime_code_table_codes(reserve_code_path)
        print(
            f"protecting IME default symbol codes from: {reserve_code_path} "
            f"({len(protected_extra_codes)} raw codepoint pairs)",
            flush=True,
        )

    lock_path = Path(args.lock)
    if not lock_path.is_absolute():
        lock_path = repo / lock_path
    if args.ignore_lock:
        print(f"ignoring existing lock: {lock_path}", flush=True)
        lock = {}
    else:
        lock = load_lock(lock_path)

    allocations, warnings = allocate_chars(
        result.used_cht,
        original_char_map,
        original_code_map,
        lock,
        prefer_t2s_slots=not args.no_prefer_t2s_slots,
        reserved_ime_codes=reserved_ime_codes,
        protected_extra_codes=protected_extra_codes,
        allow_old_non_ime_spare=args.allow_old_non_ime_spare,
    )
    print(f"  allocations: {len(allocations)}")

    write_text_reports(out, result, allocations, warnings, tasks, task_counts)
    if reserved_ime_codes:
        allocated_by_code = {item.code: item for item in allocations}
        with (out / "ime_reserved_codes.tsv").open("w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter="\t", lineterminator="\n")
            writer.writerow(["code", "base_char", "ime_refs", "allocated_char", "allocation_source", "note"])
            for code in sorted(reserved_ime_codes):
                base = original_code_map.get(code)
                alloc = allocated_by_code.get(code)
                writer.writerow([
                    f"${code:04X}",
                    base.char if base else "",
                    reserved_ime_counts.get(code, 0),
                    alloc.char if alloc else "",
                    alloc.source if alloc else "reserved-only",
                    alloc.note if alloc else "kept as original glyph if DFS bins were restored from base before generation",
                ])
        (out / "ime_reserved_summary.txt").write_text(
            "IME reserve summary\n"
            f"unique_codepoints={len(reserved_ime_codes)}\n"
            f"candidate_refs={sum(reserved_ime_counts.values())}\n"
            f"labels={len(reserved_ime_labels)}\n"
            "new CHT-only glyph allocations avoid these codepoints. Exact/t2s-compatible allocations may reuse them.\n",
            encoding="utf-8",
        )
    write_lock(out / "charmap_cht.lock.tsv", allocations, lock)
    write_charmap_txt(out / "charmap.txt", original_lines, allocations)
    write_charmap_asm(out / "charmap.asm", repo, allocations)

    if not args.no_font:
        print("generating DFS font bins from Fusion Pixel Font...")
        preview, font_source = generate_font_assets(repo, out, allocations, args)
        print(f"  preview: {preview}")
        print(f"  font: {font_source.font_path}")

    if args.apply_xlsx:
        apply_xlsx_conversion(repo, result)

    if args.apply_assets:
        apply_assets(repo, out, apply_charmap_asm=args.apply_charmap_asm)
        write_lock(lock_path, allocations, lock)
        print(f"updated lock: {lock_path}")

    print(f"done. reports/assets: {out}")
    # Report newly allocated private holes and any remaining collisions against original symbols.
    with (out / "private_hole_allocations.tsv").open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter="\t", lineterminator="\n")
        writer.writerow(["char", "code", "source", "note"])
        for item in sorted(allocations, key=lambda a: (a.code, a.char)):
            if item.code not in original_code_map:
                writer.writerow([item.char, item.code_hex, item.source, item.note])
    with (out / "legacy_collision_audit.tsv").open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter="\t", lineterminator="\n")
        writer.writerow(["code", "base_char", "allocated_char", "source", "risk"])
        for item in sorted(allocations, key=lambda a: (a.code, a.char)):
            base = original_code_map.get(item.code)
            if not base or base.char == item.char:
                continue
            ok, reason = code_is_compatible_with_original_char(item.char, item.code, original_code_map, None)
            risk = "compatible" if ok else ("symbol-collision" if original_entry_is_preserved_special(base) else "legacy-cjk-reused")
            writer.writerow([item.code_hex, base.char, item.char, item.source, risk])

    print("review summary.txt, build_text_manifest.tsv, extract_task_summary.tsv, charmap_allocation.tsv, ime_reserved_codes.tsv, private_hole_allocations.tsv, legacy_collision_audit.tsv, charmap.txt, charmap_asm_note.txt, and dfs_font_preview.png before applying.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
