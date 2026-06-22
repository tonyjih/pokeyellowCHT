#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Compare trailing @ counts in src/xlsx/data.xlsx between a git revision and the working tree.

Focuses on variable-length list sheets where extra trailing @ breaks runtime indexing:
Items, Moves, Townnames, Trainers, Patches, and Main except fixed-width trade nickname rows.
"""

from __future__ import annotations

import argparse
import subprocess
import tempfile
from pathlib import Path
from typing import Set

from openpyxl import load_workbook

VERSION = "compare-data-xlsx-at-padding-2026-06-18-v1"

VARIABLE_SHEETS = {"Items", "Moves", "Townnames", "Trainers", "Patches"}
DEFAULT_MAIN_KEEP_RANGES = "297-303"


def parse_ranges(spec: str) -> Set[int]:
    out: Set[int] = set()
    if not spec:
        return out
    for part in spec.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            a, b = part.split("-", 1)
            x, y = int(a), int(b)
            if y < x:
                x, y = y, x
            out.update(range(x, y + 1))
        else:
            out.add(int(part))
    return out


def trailing_at_count(v) -> int:
    if not isinstance(v, str):
        return 0
    n = 0
    for ch in reversed(v):
        if ch == "@":
            n += 1
        else:
            break
    return n


def git_show(repo: Path, rev: str, path: str, out: Path) -> bool:
    p = subprocess.run(
        ["git", "-C", str(repo), "show", f"{rev}:{path}"],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )
    if p.returncode != 0:
        out.write_bytes(b"")
        return False
    out.write_bytes(p.stdout)
    return True


def main() -> None:
    ap = argparse.ArgumentParser(description="Compare trailing @ padding in data.xlsx variable-length tables.")
    ap.add_argument("--version", action="store_true")
    ap.add_argument("--repo", default=".")
    ap.add_argument("--base", default="HEAD~1", help="git revision to compare against")
    ap.add_argument("--xlsx", default="src/xlsx/data.xlsx")
    ap.add_argument("--main-keep-rows", default=DEFAULT_MAIN_KEEP_RANGES)
    ap.add_argument("--out-dir", default="reports/xlsx_at_padding_compare")
    args = ap.parse_args()

    if args.version:
        print(VERSION)
        return

    repo = Path(args.repo).resolve()
    current_path = repo / args.xlsx
    out_dir = repo / args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    keep_main = parse_ranges(args.main_keep_rows)

    with tempfile.TemporaryDirectory() as td:
        base_path = Path(td) / "base.xlsx"
        ok = git_show(repo, args.base, args.xlsx, base_path)
        if not ok:
            raise SystemExit(f"git show failed: {args.base}:{args.xlsx}")

        wb_base = load_workbook(base_path, read_only=True, data_only=False)
        wb_cur = load_workbook(current_path, read_only=True, data_only=False)

        rows = []
        for sheet in sorted((set(wb_base.sheetnames) | set(wb_cur.sheetnames))):
            if sheet not in VARIABLE_SHEETS and sheet != "Main":
                continue
            if sheet not in wb_base.sheetnames or sheet not in wb_cur.sheetnames:
                continue
            ws_b = wb_base[sheet]
            ws_c = wb_cur[sheet]
            max_row = max(ws_b.max_row or 0, ws_c.max_row or 0)
            for row in range(1, max_row + 1):
                if sheet == "Main" and row in keep_main:
                    continue
                col = 3
                b = ws_b.cell(row=row, column=col).value
                c = ws_c.cell(row=row, column=col).value
                tb = trailing_at_count(b)
                tc = trailing_at_count(c)
                if tb != tc or (isinstance(c, str) and tc > 1):
                    status = "increased" if tc > tb else "decreased" if tc < tb else "same_gt1"
                    rows.append((sheet, row, col, tb, tc, status, "" if b is None else str(b), "" if c is None else str(c)))

    report = out_dir / "compare.tsv"
    with report.open("w", encoding="utf-8", newline="") as f:
        f.write("sheet\trow\tcol\tbase_trailing_at\tcurrent_trailing_at\tstatus\tbase\tcurrent\n")
        for r in rows:
            f.write("\t".join(map(str, r)) + "\n")

    summary = out_dir / "summary.txt"
    inc = sum(1 for r in rows if r[5] == "increased")
    gt1 = sum(1 for r in rows if r[4] > 1)
    with summary.open("w", encoding="utf-8") as f:
        f.write(f"{VERSION}\n")
        f.write(f"repo={repo}\n")
        f.write(f"base={args.base}\n")
        f.write(f"xlsx={current_path}\n")
        f.write(f"rows_reported={len(rows)}\n")
        f.write(f"increased={inc}\n")
        f.write(f"current_gt1={gt1}\n")
        f.write(f"report={report.relative_to(repo)}\n")

    print(VERSION)
    print(f"rows_reported: {len(rows)}")
    print(f"increased: {inc}")
    print(f"current_gt1: {gt1}")
    print(f"wrote {report}")
    print(f"wrote {summary}")


if __name__ == "__main__":
    main()
