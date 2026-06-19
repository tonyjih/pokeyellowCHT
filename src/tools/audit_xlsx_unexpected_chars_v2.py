#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Audit xlsx cells containing suspicious/unexpected characters.

Robust version:
- skips Excel temp/lock files such as ~$data.xlsx
- verifies xlsx files are valid zip containers before openpyxl reads them
- records skipped/corrupt files in skipped_xlsx.tsv instead of crashing

Usage:
  python3 audit_xlsx_unexpected_chars.py --repo . --chars 'をんゥェ'
  python3 audit_xlsx_unexpected_chars.py --repo . --base-repo /tmp/pokeyellowCHS_base --chars 'をんゥェ'
"""
from __future__ import annotations

import argparse
import csv
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

VERSION = "xlsx-unexpected-char-audit-2026-06-18-v2"

@dataclass(frozen=True)
class Hit:
    root: str
    file: str
    sheet: str
    cell: str
    chars: str
    value: str

@dataclass(frozen=True)
class Skipped:
    root: str
    file: str
    reason: str
    detail: str


def iter_xlsx(root: Path, include_temp: bool = False, use_xlsx_list: bool = False) -> Iterable[Path]:
    xlsx_dir = root / "src" / "xlsx"
    if not xlsx_dir.exists():
        raise SystemExit(f"xlsx directory not found: {xlsx_dir}")

    if use_xlsx_list:
        list_path = xlsx_dir / "xlsxList.txt"
        if not list_path.exists():
            raise SystemExit(f"xlsxList.txt not found: {list_path}")
        names = []
        for line in list_path.read_text(encoding="utf-8", errors="replace").splitlines():
            s = line.strip()
            if not s or s.startswith("#") or s.startswith(";"):
                continue
            names.append(s)
        for name in names:
            p = xlsx_dir / name
            if p.exists():
                yield p
        return

    for p in sorted(xlsx_dir.glob("*.xlsx")):
        name = p.name
        if not include_temp and (name.startswith("~$") or name.startswith(".~") or name.startswith(".")):
            continue
        yield p


def scan_root(root: Path, chars: str, label: str, include_temp: bool = False, use_xlsx_list: bool = False) -> tuple[list[Hit], list[Skipped]]:
    hits: list[Hit] = []
    skipped: list[Skipped] = []

    for path in iter_xlsx(root, include_temp=include_temp, use_xlsx_list=use_xlsx_list):
        rel = str(path.relative_to(root))
        try:
            # Excel lock files and half-written files often have .xlsx suffix but are not valid zip files.
            if not zipfile.is_zipfile(path):
                skipped.append(Skipped(label, rel, "not_zipfile", f"size={path.stat().st_size}"))
                continue
            wb = load_workbook(path, read_only=True, data_only=False)
        except (OSError, zipfile.BadZipFile, InvalidFileException, KeyError, ValueError) as e:
            skipped.append(Skipped(label, rel, type(e).__name__, str(e)))
            continue

        try:
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        v = cell.value
                        if v is None:
                            continue
                        s = str(v)
                        found = "".join(ch for ch in chars if ch in s)
                        if found:
                            hits.append(Hit(label, rel, ws.title, cell.coordinate, found, s))
        finally:
            wb.close()

    return hits, skipped


def key_without_root(hit: Hit) -> tuple[str, str, str, str]:
    return (hit.file, hit.sheet, hit.cell, hit.value)


def write_hits_tsv(path: Path, hits: list[Hit]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter="\t", lineterminator="\n")
        w.writerow(["root", "file", "sheet", "cell", "chars", "value"])
        for h in hits:
            w.writerow([h.root, h.file, h.sheet, h.cell, h.chars, h.value])


def write_skipped_tsv(path: Path, rows: list[Skipped]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter="\t", lineterminator="\n")
        w.writerow(["root", "file", "reason", "detail"])
        for r in rows:
            w.writerow([r.root, r.file, r.reason, r.detail])


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--repo", default=".", help="current repo root")
    ap.add_argument("--base-repo", default=None, help="optional clean/base repo root to compare")
    ap.add_argument("--chars", default="をんゥェ", help="characters to search")
    ap.add_argument("--out-dir", default="reports/xlsx_unexpected_chars")
    ap.add_argument("--include-temp", action="store_true", help="include Excel temp/lock files such as ~$*.xlsx")
    ap.add_argument("--use-xlsx-list", action="store_true", help="scan only files listed in src/xlsx/xlsxList.txt")
    ap.add_argument("--version", action="store_true")
    args = ap.parse_args()

    if args.version:
        print(VERSION)
        return

    repo = Path(args.repo).resolve()
    out = repo / args.out_dir
    current_hits, skipped = scan_root(repo, args.chars, "current", include_temp=args.include_temp, use_xlsx_list=args.use_xlsx_list)
    write_hits_tsv(out / "current_hits.tsv", current_hits)

    print(VERSION)
    print(f"chars={args.chars}")
    print(f"current hits={len(current_hits)}")
    print(f"current skipped={len(skipped)}")
    print(f"current report={out / 'current_hits.tsv'}")

    if args.base_repo:
        base = Path(args.base_repo).resolve()
        base_hits, base_skipped = scan_root(base, args.chars, "base", include_temp=args.include_temp, use_xlsx_list=args.use_xlsx_list)
        skipped.extend(base_skipped)
        write_hits_tsv(out / "base_hits.tsv", base_hits)
        base_keys = {key_without_root(h) for h in base_hits}
        cur_keys = {key_without_root(h) for h in current_hits}
        current_only = [h for h in current_hits if key_without_root(h) not in base_keys]
        base_only = [h for h in base_hits if key_without_root(h) not in cur_keys]
        write_hits_tsv(out / "current_only.tsv", current_only)
        write_hits_tsv(out / "base_only.tsv", base_only)
        print(f"base hits={len(base_hits)}")
        print(f"base skipped={len(base_skipped)}")
        print(f"current_only={len(current_only)}")
        print(f"base_only={len(base_only)}")
        print(f"compare reports={out}")

    write_skipped_tsv(out / "skipped_xlsx.tsv", skipped)
    if skipped:
        print(f"skipped report={out / 'skipped_xlsx.tsv'}")

if __name__ == "__main__":
    main()
