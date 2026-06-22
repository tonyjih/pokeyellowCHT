from openpyxl import load_workbook
import sys
from openpyxl.styles import Color, PatternFill, Font, Border
import openpyxl
from enum import Enum
import os

warningFill = PatternFill(start_color='0000FFFF',
                   end_color='0000FFFF',
                   fill_type='solid')
noFill = openpyxl.styles.PatternFill(fill_type=None)
textReplacement = {'#MON':'#','#':'寶可夢','&':'訓練家'}
placer = {'<PLAYER>':'<玩家>','<RIVAL>':'<劲敌>','<USER>':'<用户>','<TARGET>':'<目标>'}
def replaceText(text,dict):
    output = removeNone(text)
    for key in dict:
        output = output.replace(key,dict[key])
    return output

class INSTName():
    text = 'text'
    text_start = 'text_start'
    line = 'line'
    cont = 'cont'
    para = 'para'
    next = 'next'
    page = 'page'
    dex = 'dex'
    text_end = 'text_end'
    done = 'done'
    prompt = 'prompt'
    text_ram = 'text_ram'
    text_decimal = 'text_decimal'
    text_bcd = 'text_bcd'
    TX_RAM = 'TX_RAM'
    TX_NUM = 'TX_NUM'
    TX_BCD = 'TX_BCD'
    PLAYER = '<PLAYER>'
    RIVAL = '<RIVAL>'

class ListMode(Enum):
    blackList = 0
    whiteList = 1

class INSTFilter:
    def __init__(self,name,upList,upListMode,downList,downListMode) -> None:
        self.name = name
        self.upList = upList
        self.upListMode = upListMode
        self.downList = downList
        self.downListMode = downListMode

class INST(Enum):
    text = INSTFilter(INSTName.text,
                      [],ListMode.blackList,
                      [INSTName.text],ListMode.blackList)
    
    text_start = INSTFilter(INSTName.text_start,
                            [INSTName.text_ram,INSTName.text_bcd,INSTName.text_decimal,INSTName.line],ListMode.whiteList,
                            [INSTName.text_start],ListMode.blackList)
    
    line = INSTFilter(INSTName.line,
                      [],ListMode.blackList,
                      [INSTName.line,INSTName.text],ListMode.blackList)
    
    para = INSTFilter(INSTName.para,
                      [],ListMode.blackList,
                      [INSTName.cont,INSTName.text],ListMode.blackList)
    
    cont = INSTFilter(INSTName.cont,
                      [],ListMode.blackList,
                      [INSTName.text],ListMode.blackList)
    
    next = INSTFilter(INSTName.next,
                      [],ListMode.blackList,
                      [INSTName.next,INSTName.page,INSTName.dex,INSTName.done,INSTName.prompt],ListMode.whiteList)
    
    page = INSTFilter(INSTName.page,
                      [INSTName.dex],ListMode.blackList,
                      [],ListMode.blackList)
    
    dex = INSTFilter(INSTName.dex,
                     [INSTName.next],ListMode.whiteList,
                     [],ListMode.whiteList)
    
    text_end = INSTFilter(INSTName.text_end,[],ListMode.blackList,[],ListMode.whiteList)
    done = INSTFilter(INSTName.done,[],ListMode.blackList,[],ListMode.whiteList)
    prompt = INSTFilter(INSTName.prompt,[],ListMode.blackList,[],ListMode.whiteList)
    
    text_ram = INSTFilter(INSTName.text_ram,
                          [INSTName.line,INSTName.para,INSTName.text,INSTName.cont],ListMode.whiteList,
                          [INSTName.text,INSTName.text_end,INSTName.text_start],ListMode.whiteList)

    text_decimal = INSTFilter(INSTName.text_decimal,
                                [INSTName.line,INSTName.para,INSTName.text,INSTName.cont],ListMode.whiteList,
                              [INSTName.text,INSTName.text_end,INSTName.text_start],ListMode.whiteList)
    
    text_bcd = INSTFilter(INSTName.text_bcd,
                          [INSTName.line,INSTName.para,INSTName.text,INSTName.cont],ListMode.whiteList,
                          [INSTName.text,INSTName.text_end,INSTName.text_start],ListMode.whiteList)

    TX_RAM = INSTFilter(INSTName.TX_RAM,
                        [],ListMode.blackList,
                        [],ListMode.blackList)
    
    TX_NUM = INSTFilter(INSTName.TX_NUM,
                        [],ListMode.blackList,
                        [],ListMode.blackList)
    
    TX_BCD = INSTFilter(INSTName.TX_BCD,
                        [],ListMode.blackList,
                        [],ListMode.blackList)
    
    # PLAYER = INSTFilter(INSTName.PLAYER,[],ListMode.blackList,[],ListMode.blackList)
    # RIVAL = INSTFilter(INSTName.RIVAL,[],ListMode.blackList,[],ListMode.blackList)

def getINSTbyName(name):
    for inst in INST:
        if inst.name == name:
            return inst._value_
    return None

def checkSingleRelationVaild(instruction,nameList,filterMode):
    inst = instruction.inst
    for name in nameList:
        if inst == name:
            return filterMode != ListMode.blackList
    return filterMode == ListMode.blackList

def checkInstRelation(instructions,sheet):
    for i in range(len(instructions)):
        filter = getINSTbyName(instructions[i].inst)
        if filter != None:
            if i > 0:
                if not checkSingleRelationVaild(instructions[i-1],filter.upList,filter.upListMode):
                    printLog(InfoType.WARNING,sheet,instructions[i], instructions[i - 1].inst + '：上方指令不匹配！' + str(filter.upListMode))
            if i < len(instructions) - 1:
                if not checkSingleRelationVaild(instructions[i+1],filter.downList,filter.downListMode):
                    printLog(InfoType.WARNING,sheet,instructions[i], instructions[i + 1].inst + '：下方指令不匹配！' + str(filter.downListMode))
class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
print(bcolors.OKGREEN)

def removeNone(text):
    if text == None:
        return ''
    return text

def isChinese(_char):
    if _char == '…' or _char == '　' or _char == '！' or _char == '？' or _char == '：' or _char == '。' or _char == '，':
        return True
    return '\u4e00' <= _char <= '\u9fa5'

def ifContainsChinese(text):
    input = removeNone(text)
    for theChar in input:
        if isChinese(theChar):
            return True
    return False

def ifOverLength(text,maxPixels):
    if text == None:
        return False
    newText = text.replace(INSTName.PLAYER,'PLAYER ').replace(INSTName.RIVAL,'RIVALN ').replace('\n','')
    if len(newText) <= 0:
        return False
    charProp = isChinese(newText[0])
    length = 1
    pixels = 0
    for i in range(1,len(newText)):
        character = newText[i]
        if character == '@':
            continue
        newProp = isChinese(character)
        if newProp == charProp:
            length += 1
        else:
            if charProp:
                chsParts = float(int(length / 2))
                if length % 2 != 0:
                    chsParts += 16.0/24.0
                pixels += chsParts * 24
            else:
                pixels += length * 8
            charProp = newProp
            length = 0
        
        if i == len(newText) - 1:
            if charProp:
                chsParts = float(int(length / 2))
                if length % 2 != 0:
                    chsParts += 16.0/24.0
                pixels += chsParts * 24
            else:
                pixels += length * 8
    # print(pixels)
    return pixels > maxPixels

halfNumChars = ['0','1','2','3','4','5','6','7','8','9']
alphabet = 'ABCDEFGHIJKLMNOPQSRTUVWXYZ!?:,.'
def ifTextContains(text,list):
    tmp = removeNone(text)
    for item in list:
        if item in tmp:
            return True
    return False

class InfoType(Enum):
    ERROR = 1
    WARNING = 2
    INFO = 3

class Instruction():
    label = ''
    inst = ''
    content = ''
    filePath = ''
    sheetTitle = ''

def printLog(type,sheet,instuction,message):
    global errors
    global warnings
    global infos
    if type == InfoType.ERROR:
        print(bcolors.FAIL)
        print('错误！')
        errors += 1
    elif type == InfoType.WARNING:
        print(bcolors.WARNING)
        print('警告！')
        warnings += 1
    else:
        print(bcolors.OKCYAN)
        print('提醒！')
        infos += 1
    print('xlsx 路径：' + xlsxListPath)
    print('sheet 标题：' + sheet.title)
    if instuction != None:
        print(instuction.label)
        print(instuction.inst + ' ' + instuction.content)
    print(message)
    print(bcolors.OKGREEN)

errors = 0
warnings = 0
infos = 0
textOldPlacerCommands=[INSTName.TX_RAM,INSTName.TX_NUM,INSTName.TX_BCD]
textPlacerCommands=[INSTName.text_ram,INSTName.text_decimal,INSTName.text_bcd]
textFinishCommands=[INSTName.text_start,INSTName.text_end,INSTName.done,INSTName.prompt,INSTName.dex]
textPlacerLastAllowedCommands=[INSTName.line,INSTName.cont,INSTName.para]
def getInstDict(col,sheet,filePath):
    outputDict = {}
    id = 2
    labelrows = []
    labels = []
    while sheet.cell(row=id, column=col).value != 'end' and id <= 10000:
        label = removeNone(sheet.cell(row=id, column=col).value)
        if col - 1 > 0:
            verVal = sheet.cell(row=id, column=col - 1).value
            if verVal != None:
                if verVal != 'RB' and verVal != 'RGB' and verVal != 'Y' and verVal != 'YEUS' and verVal != 'YEJP':
                    printLog(InfoType.ERROR,sheet,None,'row '+ str(id) + '\ncol ' + str(col) + '\n的版本符号有误！ver:'+verVal)
        if ':' in label:
            labelrows.append(id)
            labels.append(label)
            if sheet.cell(row=id, column=col + 1).value != None:
                 printLog(InfoType.WARNING,sheet,None,label + '\ncol ' + str(col) + '\n旁边有多余内容！')
        id += 1
    labelrows.append(id) 
    if sheet.cell(row=id, column=col + 1).value != None:
        printLog(InfoType.WARNING,sheet,None,'row '+ str(id) + '\ncol ' + str(col) + '\n旁边有多余内容！')
    if id >= 10000:
        printLog(InfoType.ERROR,sheet,None,sheet.title + '\ncol ' + str(col) + '\n找不到 end 结束符号！')

    for i in range(len(labelrows) - 1):
        instructions = []
        for j in range(labelrows[i],labelrows[i + 1]):
            inst = sheet.cell(row=j, column=col + 1).value
            content = sheet.cell(row=j, column=col + 2).value
            if removeNone(inst) == '' and removeNone(content) != '' and not 'mailto:' in removeNone(content):
                printLog(InfoType.ERROR,sheet,None,labels[i] +'\n' + content + '\n文本内容没有指令！')
            if sheet.cell(row=j, column=col + 3).value == 'MACRO':
                continue
            if inst != None and not ';' in inst:
                newInstructions = Instruction()
                newInstructions.label = labels[i]
                newInstructions.inst = inst
                newInstructions.content = removeNone(content)
                newInstructions.sheetTitle = sheet.title
                newInstructions.filePath = filePath
                instructions.append(newInstructions)
        #         if ifTextContains(inst,textOldPlacerCommands):
        #             printLog(InfoType.ERROR,sheet,newInstructions,'有非法内容！')
        # if len(instructions) == 0:
        #     printLog(InfoType.INFO,sheet,None,labels[i] + '\n col' + str(col) + '\n标签为空白指令')
        outputDict[labels[i]] = instructions
    
    # for key in outputDict:
    #     print('key: ' + key)
    #     instuctions = outputDict[key]
    #     for instuction in instuctions:
    #         print(instuction.label)
    #         print(instuction.inst + ' ' + instuction.content)
    #     print()
    return outputDict

textStarterCommands=[INSTName.text,INSTName.text_start,INSTName.text_ram,INSTName.text_decimal,INSTName.text_bcd,'text $4c,','text "<_CONT>@"','vc_patch Change_link_closed_inactivity_message ']
textStarterCommands2=[INSTName.text,INSTName.para,INSTName.line,INSTName.cont,INSTName.text_ram,INSTName.text_decimal,INSTName.text_bcd,INSTName.next,INSTName.page]
textEndingCommands=[INSTName.done,INSTName.prompt,INSTName.dex,INSTName.text_end]
textInstCommands=[INSTName.text_ram,INSTName.text_decimal,INSTName.text_bcd]


def ifTextIsInList(text,commands):
    for command in commands:
        if text == command:
            return True
    return False

def checkDictValid(instDict,sheet):
    for key in instDict:
        instructions = instDict[key]

        checkInstRelation(instructions,sheet)

        if len(instructions) > 0:
            # 检查开头和结尾符号
            if not ifTextIsInList(instructions[0].inst,textStarterCommands):
                printLog(InfoType.ERROR,sheet,instructions[0],'开头不合法！')
            if not ifTextIsInList(instructions[-1].inst,textEndingCommands):
                if not '@@' in instructions[-1].content:
                    printLog(InfoType.ERROR,sheet,instructions[-1],'结尾不合法！')

            # 检查战斗后文本长度
            if 'EndBattleText' in instructions[0].label:
                lengthchk = replaceText(instructions[0].content,textReplacement)
                if not ifOverLength(lengthchk,1*8):
                    printLog(InfoType.WARNING,sheet,instructions[0],'战斗后文本可能太短！')
                if ifOverLength(lengthchk,12*8):
                    printLog(InfoType.WARNING,sheet,instructions[0],'战斗后文本可能太长！')
        
        for i in range(0,len(instructions)):
            instruction = instructions[i]

            if ifTextIsInList(instruction.inst,textFinishCommands):
                if removeNone(instruction.content) != '':
                    printLog(InfoType.ERROR,sheet,instruction,'终止指令有多余文本！')
            # 检查一般文本长度
            if not ifTextIsInList(instruction.inst,textPlacerCommands):
                lengthchk = replaceText(instruction.content,textReplacement)
                if ifOverLength(lengthchk,18 * 8):
                    printLog(InfoType.WARNING,sheet,instruction,'文本可能太长！')

            # 检查 textPlacerCommands=[INSTName.text_ram,INSTName.text_decimal,INSTName.text_bcd] 是否合法
            if i > 0:
                if ifTextIsInList(instruction.inst,textPlacerCommands):
                    if len(instructions[i - 1].content) > 0:
                        if instructions[i - 1].content[-1] != '@':
                            printLog(InfoType.ERROR,sheet,instruction,'上一条结尾不是 @ 符号！')
                        # if not ifTextIsInList(instructions[i - 1].inst,textPlacerLastAllowedCommands):
                        #     printLog(InfoType.ERROR,sheet,instruction,instructions[i - 1].inst +'\n上一条结尾指令有误！')
                    else:
                        printLog(InfoType.ERROR,sheet,instruction,'上一条结尾不是 @ 符号！')

            # 检查是否是老文本
            if '@@' in instruction.content:
                printLog(InfoType.INFO,sheet,instruction,'发现旧工程文本@@！')
            
            if ifTextIsInList(instruction.inst,textStarterCommands2) and instruction.content == '':
                printLog(InfoType.INFO,sheet,instruction,'发现空白内容！')
            
            #其他检查
            if ifTextContains(instruction.content,halfNumChars) and not ifTextIsInList(instruction.inst,textPlacerCommands):
                printLog(InfoType.INFO,sheet,instruction,'发现半角数字！')
            newText = replaceText(instruction.content.upper(),placer)
            if ifTextContains(newText,alphabet) and not ifTextIsInList(instruction.inst,textPlacerCommands):
                printLog(InfoType.INFO,sheet,instruction,newText + '\n发现英文符号！')

def getCountInfoFrom(instructions):
    count = [0,0,0,0,0,0,0,0,0,0]
    for instruction in instructions:
        for i in range(len(textInstCommands)):
            if instruction.inst == textInstCommands[i]:
                count[i] += 1
    return (count[0],count[1],count[2])

def compareDicts(origDict,transDict,sheet):
    for key in origDict:
        if not key in transDict:
            printLog(InfoType.ERROR,sheet,None,key + '：标签不存在！')
        else:
            origInstructions = origDict[key]
            transInstructions = transDict[key]
            # 检查结尾是否一致
            if len(origInstructions) > 0 and len(transInstructions) > 0:
                inst1 = origInstructions[-1].inst.replace(INSTName.text_end,INSTName.text)
                inst2 = transInstructions[-1].inst.replace(INSTName.text_end,INSTName.text)
                # print(origInstructions[0].label)
                # print(inst1)
                # print(inst2)
                # print()
                if inst1 != inst2:
                     printLog(InfoType.ERROR,sheet,origInstructions[-1],'结尾指令不一致！')
            else:
                printLog(InfoType.INFO,sheet,None,key +'：标签为空白指令')

            # 检查[INSTName.text_ram,INSTName.text_decimal,INSTName.text_bcd]数量是否一致
            origReplacerCount = getCountInfoFrom(origInstructions)
            transReplacerCount = getCountInfoFrom(transInstructions)
            if origReplacerCount != transReplacerCount:
                printLog(InfoType.ERROR,sheet,origInstructions[0],'文本替代控制符号不一致！')

xlsxListPath = sys.argv[1]
colOrig = int(sys.argv[2])
colTrans = int(sys.argv[3])
bypassFileCheck = int(sys.argv[4])


jsonText = "["
def exportJSON(dict,filePath):
    global jsonText
    for key in dict:
        instructions = dict[key]
        header = '{\"Sentence\":{\"note\":\"\",\"state\":\"Unmarked\",\"text\":\"'
        body = ''
        for instruction in instructions:
            if removeNone(instruction.content) != "":
                body += instruction.content + '\\n'
        footer = '\",\"tag\":\"'+ instruction.label + '\\n' + filePath.split('/')[-1] +'\"}},'
        jsonText += header + body + footer

print(xlsxListPath +'：检查 xlsx 合法性...')
print()



wb = load_workbook(filename = xlsxListPath)
for sheet in wb._sheets:
    outputPath = sheet.cell(row=1, column=1).value
    if not os.path.isfile(outputPath):
        if bypassFileCheck == 0:
            printLog(InfoType.ERROR,sheet,None,outputPath + '\n文件不存在！')
    origDict = getInstDict(colOrig,sheet,xlsxListPath)
    TransDict = getInstDict(colTrans,sheet,xlsxListPath)
    checkDictValid(TransDict,sheet)
    compareDicts(origDict,TransDict,sheet)
    exportJSON(TransDict,xlsxListPath)

jsonText += ']'
# print(jsonText)
print('检查结束')
print(bcolors.FAIL)
print('发现错误：'+ str(errors))
print(bcolors.WARNING)
print('发现警告：'+ str(warnings))
print(bcolors.OKCYAN)
print('发现提醒：'+ str(infos))


