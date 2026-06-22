#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Rebase cht_review_candidates_v2.tsv onto the current xlsx cell contents.

Purpose:
- Older candidate TSVs were generated from build-visible extracted text.
- Some build-visible rows include a synthetic one-character trailing "@" terminator that
  _importTextData.py appends during build.
- After build_cht_text_assets v15, xlsx cells should NOT store that synthetic "@".
- This tool produces a new candidate TSV whose current/suggested/original fields match
  the xlsx storage layer, not the build-visible layer.

It does not modify any xlsx file.
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Dict, List, Tuple

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:
    raise SystemExit("error: openpyxl is required. Install with: sudo apt install python3-openpyxl") from exc

VERSION = "rebase-cht-review-candidates-to-xlsx-2026-06-18-v1"


def strip_exactly_one_trailing_at(s: str) -> str | None:
    if not isinstance(s, str):
        return None
    if not s.endswith("@"):
        return None
    if s.endswith("@@"):
        return None
    return s[:-1]


def get_cell_value(repo: Path, rel_file: str, sheet: str, row: int, col: int, cache: Dict[str, object]) -> str | None:
    rel_file = rel_file.replace("\\", "/")
    path = repo / rel_file
    if not path.exists():
        return None
    if rel_file not in cache:
        cache[rel_file] = load_workbook(path, read_only=True, data_only=False)
    wb = cache[rel_file]
    if sheet not in wb.sheetnames:
        return None
    v = wb[sheet].cell(row=row, column=col).value
    return "" if v is None else str(v)


def rebase_row(row: Dict[str, str], cell: str | None) -> Tuple[Dict[str, str], str]:
    out = dict(row)

    current = out.get("current", "")
    suggested = out.get("suggested", "")
    original = out.get("original", "")

    cur_strip = strip_exactly_one_trailing_at(current)
    sug_strip = strip_exactly_one_trailing_at(suggested)
    orig_strip = strip_exactly_one_trailing_at(original)

    if cell is None:
        return out, "missing_target"

    # Exact storage-layer candidate already.
    if cell == current:
        return out, "exact_current"
    if cell == suggested:
        return out, "exact_suggested_already"

    # Candidate was generated from build-visible synthetic terminator.
    if cur_strip is not None and sug_strip is not None:
        if cell == cur_strip:
            out["current"] = cur_strip
            out["suggested"] = sug_strip
            if orig_strip is not None:
                out["original"] = orig_strip
            return out, "rebased_synthetic_at_current"
        if cell == sug_strip:
            out["current"] = cur_strip
            out["suggested"] = sug_strip
            if orig_strip is not None:
                out["original"] = orig_strip
            return out, "rebased_synthetic_at_suggested_already"

    # Keep row unchanged. The apply tool will report mismatch if user selects it.
    return out, "unchanged_mismatch"


def main() -> int:
    ap = argparse.ArgumentParser(description="Rebase CHT review candidates onto current xlsx storage text.")
    ap.add_argument("--version", action="store_true")
    ap.add_argument("--repo", default=".")
    ap.add_argument("--candidates", default="reports/cht_review_candidates_v2.tsv")
    ap.add_argument("--out", default="reports/cht_review_candidates_rebased.tsv")
    ap.add_argument("--report", default="reports/cht_review_candidates_rebased_report.tsv")
    args = ap.parse_args()

    if args.version:
        print(VERSION)
        return 0

    repo = Path(args.repo).resolve()
    in_path = Path(args.candidates)
    if not in_path.is_absolute():
        in_path = repo / in_path
    out_path = Path(args.out)
    if not out_path.is_absolute():
        out_path = repo / out_path
    report_path = Path(args.report)
    if not report_path.is_absolute():
        report_path = repo / report_path

    if not in_path.exists():
        raise SystemExit(f"error: candidates not found: {in_path}")

    wb_cache: Dict[str, object] = {}

    with in_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter="\t")
        if not reader.fieldnames:
            raise SystemExit("error: empty candidate TSV")
        fieldnames = list(reader.fieldnames)
        rows = list(reader)

    out_rows: List[Dict[str, str]] = []
    report_rows: List[Dict[str, str]] = []
    counts: Dict[str, int] = {}

    for idx, row in enumerate(rows, start=2):
        try:
            r = int(row.get("row", "0"))
            c = int(row.get("col", "0"))
        except ValueError:
            cell = None
        else:
            cell = get_cell_value(repo, row.get("file", ""), row.get("sheet", ""), r, c, wb_cache)

        new_row, status = rebase_row(row, cell)
        counts[status] = counts.get(status, 0) + 1
        out_rows.append(new_row)
        report_rows.append({
            "line": str(idx),
            "status": status,
            "file": row.get("file", ""),
            "sheet": row.get("sheet", ""),
            "row": row.get("row", ""),
            "col": row.get("col", ""),
            "cell": "" if cell is None else cell,
            "old_current": row.get("current", ""),
            "new_current": new_row.get("current", ""),
            "old_suggested": row.get("suggested", ""),
            "new_suggested": new_row.get("suggested", ""),
        })

    out_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.parent.mkdir(parents=True, exist_ok=True)

    with out_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, delimiter="\t", lineterminator="\n")
        w.writeheader()
        for row in out_rows:
            w.writerow(row)

    with report_path.open("w", encoding="utf-8", newline="") as f:
        rf = ["line", "status", "file", "sheet", "row", "col", "cell", "old_current", "new_current", "old_suggested", "new_suggested"]
        w = csv.DictWriter(f, fieldnames=rf, delimiter="\t", lineterminator="\n")
        w.writeheader()
        for row in report_rows:
            w.writerow(row)

    print(VERSION)
    print(f"input: {in_path}")
    print(f"output: {out_path}")
    print(f"report: {report_path}")
    for k in sorted(counts):
        print(f"{k}: {counts[k]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
