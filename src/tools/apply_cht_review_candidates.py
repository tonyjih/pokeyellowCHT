#!/usr/bin/env python3
# Apply selected CHT review candidates to xlsx files.
# v2: understands synthetic one-character trailing @ terminators produced by extracted text reports.
#
# Background:
# Some candidates were generated from extracted/build-visible text where _importTextData.py
# appends one terminator "@" at build time. The actual xlsx cell must NOT contain that
# synthetic terminator. v2 therefore accepts:
#   xlsx cell == candidate.current without one trailing @
# and writes:
#   candidate.suggested without one trailing @
# but only when both current and suggested end in exactly one @.

from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:
    print("error: openpyxl is required. Install with: sudo apt install python3-openpyxl", file=sys.stderr)
    raise SystemExit(2) from exc

VERSION = "apply-cht-review-candidates-2026-06-18-v2"


@dataclass
class Candidate:
    index: int
    priority: str
    category: str
    safe: str
    file: str
    sheet: str
    row: int
    col: int
    original: str
    current: str
    suggested: str
    reason: str


def read_candidates(path: Path) -> List[Candidate]:
    rows: List[Candidate] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter="\t")
        required = ["priority", "category", "safe_to_auto_apply", "file", "sheet", "row", "col", "original", "current", "suggested", "reason"]
        missing = [c for c in required if c not in (reader.fieldnames or [])]
        if missing:
            raise SystemExit(f"error: candidate TSV missing columns: {', '.join(missing)}")
        for i, r in enumerate(reader, start=2):
            try:
                row = int(r["row"])
                col = int(r["col"])
            except ValueError as exc:
                raise SystemExit(f"error: invalid row/col at TSV line {i}") from exc
            rows.append(Candidate(
                index=i,
                priority=(r["priority"] or "").strip(),
                category=(r["category"] or "").strip(),
                safe=(r["safe_to_auto_apply"] or "").strip(),
                file=(r["file"] or "").strip().replace("\\", "/"),
                sheet=(r["sheet"] or "").strip(),
                row=row,
                col=col,
                original=r["original"] or "",
                current=r["current"] or "",
                suggested=r["suggested"] or "",
                reason=r["reason"] or "",
            ))
    return rows


def norm_set(values: List[str] | None) -> set[str]:
    if not values:
        return set()
    out: set[str] = set()
    for v in values:
        for part in v.split(","):
            p = part.strip().lower()
            if p:
                out.add(p)
    return out


def selected(c: Candidate, priorities: set[str], safes: set[str], categories: set[str]) -> bool:
    if priorities and c.priority.lower() not in priorities:
        return False
    if safes and c.safe.lower() not in safes:
        return False
    if categories and c.category.lower() not in categories:
        return False
    return True


def strip_one_synthetic_at(s: str) -> str | None:
    """Return s without exactly one trailing @; reject strings ending with @@."""
    if not s.endswith("@"):
        return None
    if s.endswith("@@"):
        return None
    return s[:-1]


def synthetic_pair(c: Candidate) -> tuple[str, str] | None:
    cur = strip_one_synthetic_at(c.current)
    sug = strip_one_synthetic_at(c.suggested)
    if cur is None or sug is None:
        return None
    return cur, sug


def main() -> int:
    ap = argparse.ArgumentParser(description="Apply selected CHT review candidates to xlsx files.")
    ap.add_argument("--version", action="store_true")
    ap.add_argument("--repo", default=".", help="Repo root. Default: .")
    ap.add_argument("--candidates", required=False, default="reports/cht_review_candidates_v2.tsv", help="Candidate TSV path. Default: reports/cht_review_candidates_v2.tsv")
    ap.add_argument("--priority", action="append", default=["high"], help="Priority filter. Can be comma-separated or repeated. Default: high")
    ap.add_argument("--safe", action="append", default=["yes"], help="safe_to_auto_apply filter. Can be comma-separated or repeated. Default: yes")
    ap.add_argument("--category", action="append", default=[], help="Optional category filter. Can be comma-separated or repeated.")
    ap.add_argument("--dry-run", action="store_true", help="Show what would change without saving files.")
    ap.add_argument("--allow-mismatch", action="store_true", help="If the target cell does not exactly match current, replace current substring when possible. Not recommended for first run.")
    ap.add_argument("--report", default="reports/cht_manual_overrides_applied.tsv", help="Output report path.")
    args = ap.parse_args()

    if args.version:
        print(VERSION)
        return 0

    repo = Path(args.repo).resolve()
    cand_path = Path(args.candidates)
    if not cand_path.is_absolute():
        cand_path = repo / cand_path
    if not cand_path.exists():
        raise SystemExit(f"error: candidate TSV not found: {cand_path}")

    priorities = norm_set(args.priority)
    safes = norm_set(args.safe)
    categories = norm_set(args.category)

    candidates = [c for c in read_candidates(cand_path) if selected(c, priorities, safes, categories)]
    if not candidates:
        print("no candidates selected")
        return 0

    grouped: Dict[str, List[Candidate]] = {}
    for c in candidates:
        grouped.setdefault(c.file, []).append(c)

    report_rows: List[Dict[str, str]] = []
    applied = skipped = mismatched = 0
    dirty_workbooks = 0

    for rel_file, cs in sorted(grouped.items()):
        xlsx_path = repo / rel_file
        if not xlsx_path.exists():
            for c in cs:
                skipped += 1
                report_rows.append({
                    "status": "missing_file",
                    "file": c.file,
                    "sheet": c.sheet,
                    "row": str(c.row),
                    "col": str(c.col),
                    "before": "",
                    "after": "",
                    "current": c.current,
                    "suggested": c.suggested,
                    "priority": c.priority,
                    "category": c.category,
                    "reason": c.reason,
                })
            continue

        wb = load_workbook(xlsx_path)
        workbook_dirty = False
        for c in cs:
            if c.sheet not in wb.sheetnames:
                skipped += 1
                report_rows.append({
                    "status": "missing_sheet",
                    "file": c.file,
                    "sheet": c.sheet,
                    "row": str(c.row),
                    "col": str(c.col),
                    "before": "",
                    "after": "",
                    "current": c.current,
                    "suggested": c.suggested,
                    "priority": c.priority,
                    "category": c.category,
                    "reason": c.reason,
                })
                continue

            ws = wb[c.sheet]
            cell = ws.cell(row=c.row, column=c.col)
            before = "" if cell.value is None else str(cell.value)
            status = ""
            after = before

            pair = synthetic_pair(c)
            synthetic_current = pair[0] if pair else None
            synthetic_suggested = pair[1] if pair else None

            if before == c.suggested:
                status = "already_applied"
                skipped += 1
            elif synthetic_suggested is not None and before == synthetic_suggested:
                status = "already_applied_synthetic_at"
                skipped += 1
            elif before == c.current:
                status = "would_apply" if args.dry_run else "applied"
                after = c.suggested
                applied += 1
                if not args.dry_run:
                    cell.value = c.suggested
                    workbook_dirty = True
            elif synthetic_current is not None and synthetic_suggested is not None and before == synthetic_current:
                status = "would_apply_synthetic_at" if args.dry_run else "applied_synthetic_at"
                after = synthetic_suggested
                applied += 1
                if not args.dry_run:
                    cell.value = synthetic_suggested
                    workbook_dirty = True
            elif args.allow_mismatch and c.current and c.current in before:
                status = "would_apply_substring" if args.dry_run else "applied_substring"
                after = before.replace(c.current, c.suggested)
                applied += 1
                if not args.dry_run:
                    cell.value = after
                    workbook_dirty = True
            elif args.allow_mismatch and synthetic_current and synthetic_suggested and synthetic_current in before:
                status = "would_apply_substring_synthetic_at" if args.dry_run else "applied_substring_synthetic_at"
                after = before.replace(synthetic_current, synthetic_suggested)
                applied += 1
                if not args.dry_run:
                    cell.value = after
                    workbook_dirty = True
            else:
                status = "mismatch"
                mismatched += 1

            report_rows.append({
                "status": status,
                "file": c.file,
                "sheet": c.sheet,
                "row": str(c.row),
                "col": str(c.col),
                "before": before,
                "after": after,
                "current": c.current,
                "suggested": c.suggested,
                "priority": c.priority,
                "category": c.category,
                "reason": c.reason,
            })

        if workbook_dirty:
            wb.save(xlsx_path)
            dirty_workbooks += 1

    report_path = Path(args.report)
    if not report_path.is_absolute():
        report_path = repo / report_path
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with report_path.open("w", encoding="utf-8", newline="") as f:
        fieldnames = ["status", "file", "sheet", "row", "col", "before", "after", "current", "suggested", "priority", "category", "reason"]
        w = csv.DictWriter(f, fieldnames=fieldnames, delimiter="\t", lineterminator="\n")
        w.writeheader()
        for r in report_rows:
            w.writerow(r)

    mode = "dry-run" if args.dry_run else "apply"
    print(f"mode: {mode}")
    print(f"selected: {len(candidates)}")
    print(f"applied_or_would_apply: {applied}")
    print(f"already_or_skipped: {skipped}")
    print(f"mismatched: {mismatched}")
    print(f"dirty_workbooks: {dirty_workbooks}")
    print(f"report: {report_path}")
    if mismatched:
        print("note: mismatched rows were not changed. Review report before using --allow-mismatch.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
